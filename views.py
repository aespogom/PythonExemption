import csv
from datetime import date, timedelta, datetime
import logging
import os
import uuid

from bootstrap3 import forms
from django.contrib import messages
from django.contrib.auth import update_session_auth_hash
from django.contrib.auth.decorators import login_required
from django.contrib.auth.forms import PasswordChangeForm, PasswordResetForm
from django.contrib.auth.tokens import PasswordResetTokenGenerator, default_token_generator
from django.contrib.humanize.templatetags.humanize import intcomma
from django.core import paginator
from django.core.files.storage import default_storage, FileSystemStorage
from django.core.mail import EmailMultiAlternatives, send_mail, BadHeaderError
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.db import IntegrityError
from django.db.models import Q,F
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render, redirect, get_object_or_404, render_to_response
from django.template import RequestContext
from django.template.loader import get_template, render_to_string
from django_tables2 import RequestConfig

from main.decorators.security import group_required, anonymous_required
from main.forms import NuevoEmpleadoForm, NuevaPlantillaForm, NuevoConocimentoForm, EmpleadoNuevoConocimientoForm, \
    NuevaPreguntaEvaluacionForm, NuevaEvaluacionForm, RespuestaForm, NuevaAutoEvaluacionForm, NuevaSeccionForm, \
    NuevaPreguntaAutoEvaluacionForm, NuevaPreguntaForm, SueldosForm, VistaSueldosForm, Sueldos_propuestosForm, \
    VistaSueldos_propuestosForm, \
    ListaDepartamentoForm, NuevaPreguntaEvaluacionResponsableForm, NuevaEvaluacionResponsableForm, MailingEmpleadosForm, \
    Sueldos_propuestos_modificarForm
from main.models import Empleados, Evaluacion, PlantillaEvaluacion, Conocimientos, Preguntas, Respuestas, \
    ref_Conocimiento, Sueldos, Sueldos_propuestos, ref_areas, ref_rol, Secciones, Tokens, Actividades
from main.tables import EmpleadosTable, EvaluacionTable, ConocimientosTable, PlantillasTable, PreguntasTable, \
    SueldosTable, Sueldos_propuestosTable, SeccionTable, EvaluacionRespTable, EvaluacionesPerfil, TokensTable, \
    ActividadesTable
from main.utilities import views_utilities
from main.utilities.carga_archivo import processFileEmpleados, processFileEvaluacion, processFileSueldos, \
    processFilePreguntas, processFileRespuestas
from main.utilities.requests_utilities import get_checkbox_value
from main.utilities.views_utilities import comprobar_usuario, comprobar_responsable, comprobar_identificador, \
    comprobar_identificador_ep, buscar_siguiente_pregunta, \
    chartEmpleado, chartResponsable, calcularProgreso, calcularSMax, calcularSMedio, \
    calcularSMin, randomStringDigits, insert_new_level, crearListaOrganigrama, chart_eval_responsable, \
    chartResponsable_propuesto
from people.settings import base
from people.settings.base import MEDIA_ROOT, EMAIL_SUFIX, RESET_TIMEOUT_DAYS
from openpyxl import Workbook

logger = logging.getLogger(__name__)
PER_PAGE = 10
PER_PAGE_EMPLEADO = 25
BASE = os.path.dirname(os.path.abspath(__file__))

departamento = ""


def handler404(request):
    return render(request, '404.html', status=404)


@login_required(login_url='/accounts/login/')
def inicioSesion(request):
    usuario = request.user.username
    
    if usuario == "admin":
        actividad = Actividades(usuario=usuario, accion='LOGIN', informacion='')
        actividad.save()
        return redirect('lista_empleados')
    else:
        empleado = Empleados.objects.get(email=usuario)
        actividad = Actividades(usuario=usuario, accion='LOGIN', informacion='')
        actividad.save()
        return redirect('empleadoPerfil', id_empleado=empleado.id_empleado)


@login_required(login_url='/accounts/login/')
@group_required('admin')
def crear_nuevo_empleado(request):
    log_access(request)
    context = dict()
    context['active_empleados'] = 'class="active"'
    context['title'] = 'Nuevo empleado'
    context['form'] = NuevoEmpleadoForm
    correo = request.POST.get('email', None)
    nombre = request.POST.get('nombre', None)
    if request.method == 'POST':
        logger.debug('%s - datos recogidos' % request.path)
        sin_responsable = get_checkbox_value(request.POST.get('sin_responsable'), False)
        f_antiguedad = request.POST.get('f_antiguedad_year', None) + '-' + request.POST.get('f_antiguedad_month',None)\
                       + '-' + request.POST.get('f_antiguedad_day', None)
        if views_utilities.comprobar_empleado(correo):
            if sin_responsable:
                Empleados(identificador=request.POST.get('identificador', None),
                          email=request.POST.get('email', None),
                          nombre=request.POST.get('nombre', None),
                          apellidos=request.POST.get('apellidos', None),
                          f_antiguedad=f_antiguedad,
                          area_empleado=ref_areas.objects.get(id_area=request.POST.get('area', None)).descripcion,
                          rol=ref_rol.objects.get(
                              id_rol=request.POST.get('rol', None)).descripcion,
                          evaluador=get_checkbox_value(request.POST.get('evaluador'), False),
                          ver_sueldos=get_checkbox_value(request.POST.get('sueldos'), False),
                          responsable=None,
                          resp_depart=get_checkbox_value(request.POST.get('res_depart'), False)).save()
            else:
                Empleados(identificador=request.POST.get('identificador', None),
                          email=request.POST.get('email', None),
                          nombre=request.POST.get('nombre', None),
                          apellidos=request.POST.get('apellidos', None),
                          f_antiguedad=f_antiguedad,
                          area_empleado=ref_areas.objects.get(
                              id_area=request.POST.get('area', None)).descripcion,
                          rol=ref_rol.objects.get(
                              id_rol=request.POST.get('rol', None)).descripcion,
                          evaluador=get_checkbox_value(request.POST.get('evaluador'), False),
                          ver_sueldos=get_checkbox_value(request.POST.get('sueldos'), False),
                          responsable=Empleados.objects.get(id_empleado=request.POST.get('responsable', None)),
                          resp_depart=get_checkbox_value(request.POST.get('res_depart'), False)).save()
                context['title'] = 'Nuevo empleado creado'
                context['nombre'] = nombre
                context['apellidos'] = request.POST.get('apellidos', None)
                context['email'] = correo
            usuario = request.user.username
            return render(request, 'empleados/nuevo_empleado_creado.html', context)
        else:
            logger.error("Empleado duplicado")
            return render(request, 'empleados/nuevo_empleado.html', context)
    else:
        logger.debug('%s - carga' % request.path)
        return render(request, 'empleados/nuevo_empleado.html', context)


@login_required(login_url='/accounts/login/')
@group_required('admin')
def ver_empleados(request):
    log_access(request)
    context = dict()
    context['title'] = 'Empleados'
    if request.method == 'GET':
        logger.debug('%s - carga de información' % request.path)

        table = EmpleadosTable(data=Empleados.objects.filter(eliminado=False).order_by('apellidos'))
        table.paginate(page=request.GET.get('page', 1), per_page=PER_PAGE_EMPLEADO)
        context['table'] = table
        RequestConfig(request).configure(table)  # to sort

    elif request.method == 'POST':
        empleados = Empleados.objects.all()
        lista_empleados = list()
        for empleado in empleados:
            contextEmail = dict()
            contrasenia = randomStringDigits(10)
            views_utilities.crear_usuario(empleado)

            # PREPARARAR EMAIL
            contextEmail['title'] = "Inicio evaluaciones"

            contextEmail['empleado'] = empleado
            correo = empleado.email
            enlace = "/empleado/perfil/" + str(empleado.id_empleado)
            # Añadir los datos al email y enviar
            contextEmail['msgEmail'] = "ya puedes entrar al proceso de evaluacion.\n Tus datos de acceso son:"
            contextEmail['enlace'] = request._current_scheme_host + enlace
            contextEmail['user'] = empleado.email
            contextEmail['cnt'] = contrasenia
            email_html = get_template('email/email.html')
            email_body = email_html.render(contextEmail)
            email = EmailMultiAlternatives('Inicio evaluaciones', email_body, "text/html", to=[correo])
            email.attach_alternative(email_body, "text/html", )
            email.send()

            lista_empleados.append(empleado.nombre + ' ' + empleado.apellidos + ' | ' + empleado.email)
            context['empleados'] = lista_empleados
            context['msg'] = "Se han enviado los correos"
        return render(request, 'empleados/usuarios_creados.html', context)

    return render(request, 'empleados/tabla_empleados.html', context)


@login_required(login_url='/accounts/login/')
@group_required('admin')
def ver_tokens(request):
    #vista para visualizar los tokens que se han generado
    log_access(request)
    context = dict()
    context['title'] = 'Tokens'
    if request.method == 'GET':
        logger.debug('%s - carga de información' % request.path)

        table = TokensTable(data=Tokens.objects.all())
        table.paginate(page=request.GET.get('page', 1), per_page=PER_PAGE_EMPLEADO)
        context['table'] = table
        RequestConfig(request).configure(table)  # to sort
    return render(request, 'empleados/tabla_tokens.html', context)


@login_required(login_url='/accounts/login/')
@group_required('admin')
def detalle_empleado(request, id_empleado):
    log_access(request)
    context = dict()
    empleado = Empleados.objects.get(id_empleado=id_empleado)
    context['id_empleado'] = empleado.id_empleado
    context['title'] = 'Detalle empleado'
    context['empleado'] = empleado.nombre + " " + empleado.apellidos
    context['correo'] = empleado.email
    context['responsable'] = empleado.responsable
    table = ConocimientosTable(data=Conocimientos.objects.filter(id_empleado=id_empleado))
    table.paginate(page=request.GET.get('page', 1), per_page=PER_PAGE)  # QUITAR COLUMNA EMLEADO
    context['table'] = table
    if request.method == 'GET':
        logger.debug('%s - carga de información' % request.path)
        table_evaluaciones = EvaluacionTable(data=Evaluacion.objects.filter(empleado=empleado))
        context['table_evaluaciones'] = table_evaluaciones
        #solo se vera si el usuario tiene permiso de verSueldo
        table_sueldos = SueldosTable(data=Sueldos.objects.filter(empleado=empleado))

        table_sueldos.paginate(page=request.GET.get('page', 1), per_page=PER_PAGE_EMPLEADO)
        context['table_sueldos'] = table_sueldos

        table_sueldos_propuestos = Sueldos_propuestosTable(data=Sueldos_propuestos.objects.filter(empleado=empleado))

        table_sueldos_propuestos.paginate(page=request.GET.get('page', 1), per_page=PER_PAGE_EMPLEADO)
        context['table_sueldos_propuestos'] = table_sueldos_propuestos
        if empleado.evaluador:
            context['esResponsable'] = True
        if empleado.ver_sueldos or empleado.responsable:
            context['verSueldo'] = True
    else:
        logger.debug('%s - subida archivo' % request.path)

    RequestConfig(request).configure(table)  # to sort
    return render(request, 'empleados/detalle_empleado.html', context)


@login_required(login_url='/accounts/login/')
@group_required('admin')
def responsable_empleados(request, id_empleado):
    log_access(request)
    context = dict()
    empleado = Empleados.objects.get(id_empleado=id_empleado)
    context['id_empleado'] = empleado.id_empleado
    context['title'] = 'Responsable empleados'
    context['empleado'] = empleado.nombre + " " + empleado.apellidos
    context['correo'] = empleado.email
    usuario = request.user.username
    if request.method == 'GET':
        logger.debug('%s - carga de información' % request.path)
        if empleado.evaluador:
            tablaEmpleados = EmpleadosTable(data=Empleados.objects.filter(responsable=empleado, eliminado=False))
            tablaEmpleados.paginate(page=request.GET.get('page', 1), per_page=5)  # QUITAR COLUMNA RESPONSABLE
            RequestConfig(request).configure(tablaEmpleados)  # to sort
            context['tablaEmpleados'] = tablaEmpleados

    else:
        logger.debug('%s - subida archivo' % request.path)

    return render(request, 'empleado/responsable_empleados.html', context)


@login_required(login_url='/accounts/login/')
@group_required('admin')
def crear_nuevo_conocimiento(request):
    log_access(request)
    context = dict()
    context['title'] = 'Nuevo conocimiento'
    new_form = NuevoConocimentoForm
    if request.method == 'GET':
        logger.debug('%s - formulario nuevo conocimiento' % request.path)
        context['form'] = new_form

    else:
        logger.debug('%s - nueva conocimiento' % request.path)
        empleado = Empleados.objects.get(id_empleado=request.POST.get('empleado', None))
        conocimiento = ref_Conocimiento.objects.get(
            id_conocimiento=request.POST.get('descripcion', None))
        if conocimiento.descripcion == "Otro":
            conocimiento = request.POST.get('otroConocimiento')
        else:
            descripcionConocimiento = ref_Conocimiento.objects.get(
                id_conocimiento=request.POST.get('descripcion', None))
            conocimiento = descripcionConocimiento.descripcion
        if not conocimiento:
            context['vacio'] = True
        else:
            usuario = request.user.username
            actividad = Actividades(usuario=usuario, accion='NUEVO CONOCIMIENTO',
                                    informacion='Conocimiento: ' + str(conocimiento) + ' Id_empleado: ' + str(request.POST.get('empleado', None)))
            actividad.save()
            if empleado.id_empleado:
                Conocimientos(descripcion=conocimiento,
                              nivel=request.POST.get('nivel', None),
                              id_empleado=empleado,
                              certificado=get_checkbox_value(request.POST.get('certificado', None))).save()
                context['msg'] = 'Conocimiento creado'
        new_form.id_empleado = empleado.id_empleado
        context['form'] = new_form

    return render(request, 'conocimientos/nuevo_conocimiento.html', context)


@login_required(login_url='/accounts/login/')
@group_required('admin')
def ver_conocimientos(request):
    log_access(request)
    context = dict()
    context['title'] = 'Conocimiento'

    if request.method == 'GET':
        logger.debug('%s - carga de información' % request.path)

        empleado = Empleados.objects.filter(eliminado=False)  # Coger empleados sin eliminar
        table = ConocimientosTable(data=Conocimientos.objects.filter(id_empleado__in=empleado).order_by('descripcion'))
        table.paginate(page=request.GET.get('page', 1), per_page=PER_PAGE)
        context['table'] = table

    RequestConfig(request).configure(table)  # to sort

    return render(request, 'conocimientos/tabla_conocimientos.html', context)


@login_required(login_url='/accounts/login/')
@group_required('admin')
def indice_evaluaciones(request):
    log_access(request)
    context = dict()
    context['title'] = 'Tipos de evaluaciones'

    logger.debug('%s - carga de pagina' % request.path)

    return render(request, 'evaluaciones/indice_evaluaciones.html', context)


@login_required(login_url='/accounts/login/')
@group_required('admin')
def crear_plantilla_autoevaluacion(request):
    log_access(request)
    context = dict()
    context['active_evaluaciones'] = 'class="active"'

    if request.method == 'POST':
        logger.debug('%s - datos recogidos' % request.path)
        PlantillaEvaluacion(descripcion=request.POST.get('descripcion', None),
                            autoevaluacion=True).save()
        context['title'] = 'Nueva plantilla creada'
        context['tipo'] = "evaluacion"
        context['descripcion'] = request.POST.get('descripcion', None)
        context['autoevaluacion'] = 'Sí'
        return render(request, 'plantillas/nueva_plantilla_creada.html', context)
    else:
        logger.debug('%s - carga' % request.path)
        context['title'] = 'Nueva Plantilla '
        context['form'] = NuevaPlantillaForm

        # select * from plantillaevaluacion where autoevaluacion = True
        context['table'] = PlantillasTable(
            data=PlantillaEvaluacion.objects.filter(autoevaluacion=True).order_by('descripcion'))
        # RequestConfig(request).configure(table)  # to sort
        return render(request, 'evaluaciones/plantilla_autoevaluacion.html', context)


@login_required(login_url='/accounts/login/')
@group_required('admin')
def nueva_autoevaluaciones(request):
    log_access(request)
    context = dict()
    context['title'] = 'Nueva autoevaluacion'

    if request.method == 'GET':
        logger.debug('%s - carga de información' % request.path)
        context['form'] = NuevaAutoEvaluacionForm
        return render(request, 'evaluaciones/nueva_autoevaluacion.html', context)
    else:
        id_plantilla = request.POST.get('id_plantilla', None)
        id_responsable = request.POST.get('responsable', None)
        id_empleado = request.POST.get('empleado', None)
        if id_plantilla and id_responsable and id_empleado:
            plantilla = PlantillaEvaluacion.objects.get(id_plantilla=id_plantilla)
            responsable = Empleados.objects.get(id_empleado=id_responsable)
            empleado = Empleados.objects.get(id_empleado=id_empleado)

            Evaluacion(id_plantilla=plantilla,
                       anio=request.POST.get('anio', None),
                       detalle=request.POST.get('detalle', None),
                       responsable=responsable,
                       empleado=empleado).save()
            context['plantilla'] = plantilla
            context['anio'] = request.POST.get('anio', None)
            context['responsable'] = responsable
            context['empleado'] = empleado
            preguntas = Preguntas.objects.filter(id_plantilla=plantilla.id_plantilla).order_by('updated')
            context['preguntas'] = [pregunta.pregunta for pregunta in preguntas]

        return render(request, 'evaluaciones/nueva_autoevaluacion_creada.html', context)


@login_required(login_url='/accounts/login/')
@group_required('admin')
def ver_autoevaluaciones(request):
    log_access(request)
    context = dict()
    context['title'] = 'Autoevaluaciones'
    context['tipoEvaluacion'] = "auto evaluacion"
    if request.method == 'GET':
        logger.debug('%s - carga de información' % request.path)
        list_ids_plantilla = PlantillaEvaluacion.objects.filter(autoevaluacion=True).values_list('id_plantilla',
                                                                                                 flat=True)
        empleado = Empleados.objects.filter(eliminado=False)  # Coger empleados sin eliminar
        evaluacion = Evaluacion.objects.filter(id_plantilla__in=list_ids_plantilla, empleado__in=empleado)
        table = EvaluacionTable(
            data=evaluacion.order_by('anio'))
        table.paginate(page=request.GET.get('page', 1), per_page=PER_PAGE)
        context['table'] = table
    RequestConfig(request).configure(table)  # to sort
    return render(request, 'evaluaciones/tabla_autoevaluaciones.html', context)


@login_required(login_url='/accounts/login/')
@group_required('admin')
def ver_evaluaciones(request):
    log_access(request)
    context = dict()
    context['title'] = 'Evaluaciones'

    if request.method == 'GET':
        logger.debug('%s - carga de información' % request.path)

    return render(request, 'evaluaciones/tabla_evaluaciones.html', context)


@login_required(login_url='/accounts/login/')
@group_required('admin')
def todas_evaluacion(request):
    log_access(request)
    context = dict()
    context['title'] = 'Evaluaciones'
    context['tipoEvaluacion'] = "evaluacion"
    if request.method == 'GET':
        logger.debug('%s - carga de información' % request.path)
        list_ids_plantilla = PlantillaEvaluacion.objects.filter(autoevaluacion=False).values_list('id_plantilla',
                                                                                                  flat=True)
        empleado = Empleados.objects.filter(eliminado=False)  # Coger empleados sin eliminar
        evaluacion = Evaluacion.objects.filter(id_plantilla__in=list_ids_plantilla,
                                               empleado__in=empleado)  # Coger la evaluacion
        table = EvaluacionTable(
            data=evaluacion.order_by('anio'))
        table.paginate(page=request.GET.get('page', 1), per_page=PER_PAGE)
        context['table'] = table
        RequestConfig(request).configure(table)  # to sort

    return render(request, 'evaluaciones/todas_evaluaciones.html', context)


@login_required(login_url='/accounts/login/')
@group_required('admin')
def nueva_evaluacion(request):
    log_access(request)
    context = dict()
    context['title'] = 'Nueva evaluacion'

    if request.method == 'GET':
        logger.debug('%s - carga de información' % request.path)
        context['form'] = NuevaEvaluacionForm
        return render(request, 'evaluaciones/nueva_evaluacion.html', context)
    else:
        id_plantilla = request.POST.get('id_plantilla', None)
        id_responsable = request.POST.get('responsable', None)
        id_empleado = request.POST.get('empleado', None)
        if id_plantilla and id_responsable and id_empleado:
            plantilla = PlantillaEvaluacion.objects.get(id_plantilla=id_plantilla)
            responsable = Empleados.objects.get(id_empleado=id_responsable)
            empleado = Empleados.objects.get(id_empleado=id_empleado)

            Evaluacion(id_plantilla=plantilla,
                       anio=request.POST.get('anio', None),
                       detalle=request.POST.get('detalle', None),
                       responsable=responsable,
                       empleado=empleado).save()
            usuario = request.user.username

            context['plantilla'] = plantilla
            context['anio'] = request.POST.get('anio', None)
            context['responsable'] = responsable
            context['empleado'] = empleado
            preguntas = Preguntas.objects.filter(id_plantilla=plantilla.id_plantilla).order_by('updated')
            context['preguntas'] = [pregunta.pregunta for pregunta in preguntas]

        return render(request, 'evaluaciones/nueva_evaluacion_creada.html', context)


@login_required(login_url='/accounts/login/')
@group_required('admin')
def detalle_evaluacion(request, id_evaluacion):
    log_access(request)
    context = dict()
    context['title'] = 'Detalle evaluacion'
    if request.method == 'GET':
        logger.debug('%s - carga de información' % request.path)
        evaluacion = Evaluacion.objects.get(id_evaluacion=id_evaluacion)
        empleado = evaluacion.empleado

        context['empleado'] = empleado
        context['responsable'] = evaluacion.responsable
        context['id_evaluacion'] = evaluacion.id_evaluacion
        context['anio'] = evaluacion.anio
        if Preguntas.objects.filter(id_plantilla=evaluacion.id_plantilla.id_plantilla).exists():
            table = PreguntasTable(
                data=Preguntas.objects.filter(id_plantilla=evaluacion.id_plantilla.id_plantilla).order_by(
                    'id_plantilla'))
            table.paginate(page=request.GET.get('page', 1), per_page=PER_PAGE)
            context['table'] = table
            evaluacion = Evaluacion.objects.get(id_evaluacion=id_evaluacion)
            enlace = "/evaluacion/preguntas/lista/" + str(id_evaluacion) + "/" + str(evaluacion.empleado.id_empleado)
            context['enlace'] = enlace

            RequestConfig(request).configure(table)  # to sort
        else:
            context['msg'] = 'Sin Preguntas'
            context['sin_datos'] = True
    return render(request, 'evaluaciones/detalle_evaluacion.html', context)


@login_required(login_url='/accounts/login/')
@group_required('admin')
def nueva_seccion_preguntas(request):
    log_access(request)
    context = dict()
    context['title'] = 'Secciones'
    new_form = NuevaSeccionForm
    context['form'] = new_form
    tableSecciones = SeccionTable(data=Secciones.objects.all())
    context['tableSecciones'] = tableSecciones

    if request.method == 'GET':
        logger.debug('%s - formulario nueva seccion' % request.path)

    else:
        try:
            logger.debug('%s - nueva seccion' % request.path)
            Secciones(descripcion=request.POST.get('seccion', None)).save()
            context['msg'] = 'Sección creada'
        except IntegrityError:
            context['msg'] = 'ERROR: La sección ya existe'

    return render(request, 'preguntas/nueva_seccion_preguntas.html', context)


@login_required(login_url='/accounts/login/')
@group_required('admin')
def nueva_pregunta_evaluacion(request):
    log_access(request)
    context = dict()
    context['title'] = 'Nueva pregunta evaluacion'
    context['link'] = 'nueva_pregunta_evaluacion'

    if request.method == 'GET':
        logger.debug('%s - formulario nueva pregunta' % request.path)
        context['form'] = NuevaPreguntaEvaluacionForm
    else:
        logger.debug('%s - nueva pregunta' % request.path)
        id_plantilla = request.POST.get('id_plantilla', None)
        if id_plantilla:
            plantilla = PlantillaEvaluacion.objects.get(id_plantilla=id_plantilla)

            Preguntas(pregunta=request.POST.get('pregunta', None),
                      id_plantilla=plantilla,
                      id_seccion=Secciones.objects.get(id_seccion=request.POST.get('id_seccion', None))).save()
            context['msg'] = 'Pregunta creada'
            new_form = NuevaPreguntaEvaluacionForm
            new_form.id_plantilla = id_plantilla
            context['form'] = new_form

    return render(request, 'preguntas/nueva_pregunta.html', context)


@login_required(login_url='/accounts/login/')
@group_required('admin')
def nueva_pregunta_autoevaluacion(request):
    log_access(request)
    context = dict()
    context['title'] = 'Nueva pregunta autoevaluacion'
    context['link'] = 'nueva_pregunta_autoevaluacion'

    if request.method == 'GET':
        logger.debug('%s - formulario nueva pregunta' % request.path)
        context['form'] = NuevaPreguntaAutoEvaluacionForm
    else:
        logger.debug('%s - nueva pregunta' % request.path)
        id_plantilla = request.POST.get('id_plantilla', None)
        if id_plantilla:
            plantilla = PlantillaEvaluacion.objects.get(id_plantilla=id_plantilla)

            Preguntas(pregunta=request.POST.get('pregunta', None),
                      id_plantilla=plantilla,
                      id_seccion=Secciones.objects.get(id_seccion=request.POST.get('id_seccion', None))).save()
            context['msg'] = 'Pregunta creada'
            new_form = NuevaPreguntaAutoEvaluacionForm
            new_form.id_plantilla = id_plantilla
            context['form'] = new_form

    return render(request, 'preguntas/nueva_pregunta.html', context)


@login_required(login_url='/accounts/login/')
@group_required('admin')
def nueva_pregunta_evaluacion_responsable(request):
    log_access(request)
    context = dict()
    context['title'] = 'Nueva pregunta evaluacion responsable'
    context['link'] = 'nueva_pregunta_responsable'
    if request.method == 'GET':
        logger.debug('%s - formulario nueva pregunta' % request.path)
        context['form'] = NuevaPreguntaEvaluacionResponsableForm
    else:
        logger.debug('%s - nueva pregunta' % request.path)
        id_plantilla = request.POST.get('id_plantilla', None)
        if id_plantilla:
            plantilla = PlantillaEvaluacion.objects.get(id_plantilla=id_plantilla)

            Preguntas(pregunta=request.POST.get('pregunta', None),
                      id_plantilla=plantilla,
                      id_seccion=Secciones.objects.get(id_seccion=request.POST.get('id_seccion', None))).save()
            context['msg'] = 'Pregunta creada'
            new_form = NuevaPreguntaEvaluacionForm
            new_form.id_plantilla = id_plantilla
            context['form'] = new_form

    return render(request, 'preguntas/nueva_pregunta.html', context)


@login_required(login_url='/accounts/login/')
@group_required('admin')
def ver_preguntas_evaluacion(request):
    log_access(request)
    context = dict()
    context['title'] = 'Preguntas'

    if request.method == 'GET':
        logger.debug('%s - carga de información' % request.path)

        table = PreguntasTable(data=Preguntas.objects.all().order_by('id_plantilla'))
        table.paginate(page=request.GET.get('page', 1), per_page=PER_PAGE)
        context['table'] = table

    RequestConfig(request).configure(table)  # to sort
    return render(request, 'preguntas/tabla_preguntas_evaluacion.html', context)


@login_required(login_url='/accounts/login/')
@group_required('admin')
def ver_plantilla(request):
    log_access(request)
    context = dict()
    context['active_plantillas'] = 'class="active"'
    context['title'] = 'Plantillas'

    if request.method == 'GET':
        logger.debug('%s - carga de información' % request.path)

        # Tabla plantillas evaluaciones
        table_plantilla_evaluaciones = PlantillasTable(
            data=PlantillaEvaluacion.objects.filter(autoevaluacion=False, de_responsable=False).order_by('updated'))
        table_plantilla_evaluaciones.paginate(page=request.GET.get('page', 1), per_page=PER_PAGE)
        table_plantilla_evaluaciones.exclude = ('autoevaluacion', 'de_responsable')
        context['table_plantilla_evaluaciones'] = table_plantilla_evaluaciones

        # Tabla plantillas autoevaluaciones
        table_plantilla_autoevaluaciones = PlantillasTable(
            data=PlantillaEvaluacion.objects.filter(autoevaluacion=True, de_responsable=False).order_by('updated'))
        table_plantilla_autoevaluaciones.paginate(page=request.GET.get('page', 1), per_page=PER_PAGE)
        table_plantilla_autoevaluaciones.exclude = ('autoevaluacion', 'de_responsable')
        context['table_plantilla_autoevaluaciones'] = table_plantilla_autoevaluaciones

        # Tabla plantillas evaluaciones responsable
        table_plantilla_responsable = PlantillasTable(
            data=PlantillaEvaluacion.objects.filter(autoevaluacion=False, de_responsable=True).order_by('updated'))
        table_plantilla_responsable.paginate(page=request.GET.get('page', 1), per_page=PER_PAGE)
        table_plantilla_responsable.exclude = ('autoevaluacion', 'de_responsable')
        context['table_plantilla_responsable'] = table_plantilla_responsable

    RequestConfig(request).configure(table_plantilla_evaluaciones)  # to sort
    RequestConfig(request).configure(table_plantilla_autoevaluaciones)  # to sort
    RequestConfig(request).configure(table_plantilla_responsable)  # to sort
    return render(request, 'plantillas/tabla_plantillas.html', context)


@login_required(login_url='/accounts/login/')
@group_required('admin')
def nueva_plantillas(request):
    log_access(request)
    context = dict()
    context['active_plantillas'] = 'class="active"'

    if request.method == 'POST':
        logger.debug('%s - datos recogidos' % request.path)
        PlantillaEvaluacion(descripcion=request.POST.get('descripcion', None),
                            autoevaluacion=False).save()
        context['title'] = 'Nueva plantilla creada'
        context['tipo'] = "evaluacion"
        context['descripcion'] = request.POST.get('descripcion', None)
        context['autoevaluacion'] = 'No'
        return render(request, 'plantillas/nueva_plantilla_creada.html', context)
    else:
        logger.debug('%s - carga' % request.path)
        context['title'] = 'Nueva plantilla'
        context['form'] = NuevaPlantillaForm
        context['table'] = PlantillasTable(
            data=PlantillaEvaluacion.objects.filter(autoevaluacion=False).order_by('descripcion'))

        return render(request, 'plantillas/plantilla_evaluacion.html', context)


@login_required(login_url='/accounts/login/')
@group_required('admin')
def evaluacion_sin_verificar(request, id_empleado):
    log_access(request)
    context = dict()
    responsable = Empleados.objects.get(id_empleado=id_empleado)
    context['id_empleado'] = responsable.id_empleado
    context['responsable'] = responsable
    context['id_responsable'] = responsable.id_empleado
    context['title'] = 'Evaluaciones sin verificar'
    if request.method == 'GET':
        logger.debug('%s - carga de información' % request.path)

        table = EvaluacionTable(
            data=Evaluacion.objects.filter(responsable=responsable, verificado=False).order_by('anio'))
        table.paginate(page=request.GET.get('page', 1), per_page=PER_PAGE)
        context['table'] = table

        context['message'] = "ARCHIVO RECIBIDO"
    elif request.method == 'POST':
        log_access(request)
        contextEmail = dict()
        context = dict()
        empleados = dict()
        contextEmail['title'] = "Nuevos Mercados"
        contextEmail['msg'] = "Detalles empleado"
        evaluaciones = Evaluacion.objects.filter(responsable=Empleados.objects.get(id_empleado=responsable.id_empleado),
                                                 verificado=False)
        # Recorrer las evaluaciones y coger el email
        for evaluacion in evaluaciones:
            correo = evaluacion.empleado.email
            empleados[evaluacion.empleado] = correo
            if evaluacion.id_plantilla.autoevaluacion:
                contextEmail['empleado'] = evaluacion.empleado
                contextEmail['msgEmail'] = "Le recordamos que tiene auto evaluaciones por hacer"
            else:
                contextEmail['empleado'] = evaluacion.responsable
                contextEmail['msgEmail'] = "Le recordamos que tiene evaluaciones por hacer"
            enlace = "/empleado/perfil/" + str(evaluacion.empleado.id_empleado)
            # Añadir los datos al email y enviar
            contextEmail['enlace'] = request._current_scheme_host + enlace
            email_html = get_template('email/email.html')
            email_body = email_html.render(contextEmail)
            email = EmailMultiAlternatives('Evaluaciones pendientes', email_body, "text/html", to=[correo])
            email.attach_alternative(email_body, "text/html", )
            email.send()

        context['empleados'] = empleados
        context['msg'] = "Se han enviado los correos"
        return render(request, 'evaluaciones/correo_enviado.html', context)
    RequestConfig(request).configure(table)  # to sort
    return render(request, 'evaluaciones/detalle_evaluacion_sin_verificar.html', context)


@login_required(login_url='/accounts/login/')
@group_required('admin')
def detalle_plantilla(request, id_plantilla):
    log_access(request)
    context = dict()
    plantilla = PlantillaEvaluacion.objects.get(id_plantilla=id_plantilla)
    context['id_plantilla'] = id_plantilla
    context['title'] = "Detalle plantilla"
    context['nombre'] = plantilla.descripcion

    if plantilla.de_responsable:
        context['tipo'] = 'De responsable'
    elif plantilla.autoevaluacion:
        context['tipo'] = 'Autoevaluacion'
    else:
        context['tipo'] = 'Evaluacion'
    table = PreguntasTable(data=Preguntas.objects.filter(id_plantilla=plantilla).order_by('id_plantilla'))
    table.paginate(page=request.GET.get('page', 1), per_page=PER_PAGE)
    context['table'] = table
    if id_plantilla:
        new_form = NuevaPreguntaForm(
            initial={"id_plantilla": PlantillaEvaluacion.objects.get(id_plantilla=id_plantilla)})
        context['msg'] = 'Pregunta creada'
        context['form'] = new_form
    if request.method == 'GET':
        context['title'] = "Detalle plantilla"
    elif request.method == 'POST':
        plantilla = PlantillaEvaluacion.objects.get(id_plantilla=id_plantilla)
        Preguntas(pregunta=request.POST.get('pregunta', None),
                  id_plantilla=plantilla,
                  id_seccion=Secciones.objects.get(id_seccion=request.POST.get('id_seccion', None))
                  ).save()
    new_form.fields['id_plantilla'].widget = forms.HiddenInput()
    RequestConfig(request).configure(table)  # to sort
    return render(request, 'plantillas/detalle_plantilla.html', context)


@login_required(login_url='/accounts/login/')
@group_required('admin')
def mailing_evaluaciones(request):
    log_access(request)
    context = dict()
    context['title'] = "Mailing empleados"

    if request.method == 'GET':
        logger.debug('GET /mailing_evaluaciones/ - carga inicial')
        context['form'] = MailingEmpleadosForm
        return render(request, 'empleados/mailing_evaluaciones.html', context)
    elif request.method == 'POST':
        logger.debug('POST /mailing_evaluaciones/ - Envio de emails')
        form = MailingEmpleadosForm(request.POST)
        if form.is_valid():
            logger.debug('POST /mailing_evaluaciones/ - Form validado y recogido correctamente')
            empleados_mailing = form.cleaned_data.get('empleados')

            logger.debug('POST - /mailing_evaluaciones/ - se van a enviar %s emails' % len(empleados_mailing))

            # PREPARARAR EMAIL
            for e_mailing in empleados_mailing:
                logger.debug('POST - /mailing_evaluaciones/ - empleado %s' % e_mailing)
                e_info = Empleados.objects.get(email=e_mailing)
                if not e_info:
                    logger.error('POST - // - No se ha encontrado empleado con email %s' % e_mailing)
                    continue
                logger.debug('POST - /mailing_evaluaciones/ - prepara email para empleado %s' % e_info.nombre)
                contextEmail = dict()
                contextEmail['title'] = "Inicio evaluaciones"

                contextEmail['empleado'] = e_info.nombre
                correo = e_info.email
                enlace = "/empleado/perfil/" + str(e_info.id_empleado)
                # Añadir los datos al email y enviar
                contextEmail['msgEmail'] = "ya puedes entrar al proceso de evaluacion.\n Tus datos de acceso son:"
                contextEmail['enlace'] = request._current_scheme_host + enlace
                contextEmail['user'] = e_info.email
                contextEmail['cnt'] = 'La que tuvieras anteriormente'
                email_html = get_template('email/email.html')
                email_body = email_html.render(contextEmail)
                email = EmailMultiAlternatives('Inicio evaluaciones', email_body, "text/html", to=[correo])
                email.attach_alternative(email_body, "text/html", )
                logger.debug('POST - /mailing_evaluaciones/ - listo email para empleado %s' % e_info.nombre)
                email.send()
                logger.debug('POST - /mailing_evaluaciones/ - email enviado a empleado %s' % e_info.nombre)

            context['msg'] = 'Se han enviado %s emails' % len(empleados_mailing)


        else:
            logger.error('POST - // - No se ha podido cargar el formulario de forma correcta')
            context['msg'] = 'Error leyendo los empleados seleccionados'

        return render(request, 'empleados/mailing_evaluaciones_enviados.html', context)


@login_required(login_url='/accounts/login/')
@group_required('admin')
def cargar_empleado(request):
    log_access(request)
    context = dict()
    context['title'] = "Carga empleado"
    context['modelo'] = 1
    context['msg2'] = "ORDEN CSV: identificador, email, nombre, apellidos, f_antiguedad(YYYY-MM-DD), area_empleado, rol," \
                      " evaluador(si=1, no=0), id_responsable, responsable_departamento(si=1, no=0), ver_sueldos(si=1, no=0)"

    if request.method == 'GET':
        logger.debug('/subida_archivo_empleado/ - carga inicial')
    elif request.method == 'POST':
        try:
            if not 'file' in request.FILES:
                context['msg'] = "ERROR: No ha seleccionado archivo."
            elif not str(request.FILES['file']).lower().endswith('.csv'):
                context['msg'] = "ERROR: No es un archivo CSV."
            else:
                try:
                    if processFileEmpleados(csv_file=request.FILES['file']):
                        context['msg'] = "Empleados creados"

                except Exception as ex:
                    context['msg'] = "Error al subir empleados\n", ex
        except AttributeError as attError:
            context['msg'] = 'ERROR: Formato incorrecto', attError
        except StopIteration as attError:
            context['msg'] = 'ERROR: El fichero está vacío', attError
        except Exception as exception:
            context['msg'] = 'ERROR: ' + type(exception).__name__ + ': ' + str(exception.__str__())
    return render(request, 'empleados/carga_empleados.html', context)


@login_required(login_url='/accounts/login/')
@group_required('admin')
def cargar_sueldos(request):
    log_access(request)
    context = dict()
    context['title'] = "Carga sueldos"
    context['modelo'] = 2
    context['msg2'] = "ORDEN CSV: id_empleado, retribucion_fija, variable_individual, variable_empresa, " \
                      "beneficios_sociales, bonus_dietas, kilometros, guardias, tipo, total, incremento, año. " \
                      "En caso de ser propuesto, el año puede ir vacío. "

    if request.method == 'GET':
        logger.debug('/subida_archivo_sueldos/ - carga inicial')
    elif request.method == 'POST':
        try:
            if not 'file' in request.FILES:
                context['msg'] = "ERROR: No ha seleccionado archivo."
            elif not str(request.FILES['file']).lower().endswith('.csv'):
                context['msg'] = "ERROR: No es un archivo CSV."
            else:
                try:
                    if processFileSueldos(csv_file=request.FILES['file']):
                        context['msg'] = "Sueldos creados exitosamente :) "

                except Exception as ex:
                    context['msg'] = "Error al subir sueldos\n", ex
        except AttributeError as attError:
            context['msg'] = 'ERROR: Formato incorrecto', attError
        except StopIteration as attError:
            context['msg'] = 'ERROR: El fichero está vacío', attError
        except Exception as exception:
            context['msg'] = 'ERROR: ' + type(exception).__name__ + ': ' + str(exception.__str__())
    return render(request, 'empleados/carga_sueldos.html', context)


@login_required(login_url='/accounts/login/')
@group_required('admin')
def aniadir_sueldo(request):
    log_access(request)
    context = dict()
    new_form = SueldosForm
    context['title'] = "Sueldos"
    context['subtitle'] = "Por favor, introduzca los datos numéricos sin puntos ni comas"
    #no lo admite segun models, da error si se intenta

    if request.method == 'GET':
        context['form'] = new_form
    elif request.method == 'POST':
        id_empleado = request.POST.get('empleado', None)
        empleado = Empleados.objects.get(id_empleado=id_empleado)
        try:
            data = Sueldos.objects.get(empleado=empleado, anio=request.POST.get('anio', None))
            if data:
                data.retribucion_fija = request.POST.get('retribucion_fija', None)
                data.varibale_individual = request.POST.get('varibale_individual', None)
                data.varibale_empresa = request.POST.get('varibale_empresa', None)
                data.beneficios_sociales = request.POST.get('beneficios_sociales', None)
                data.bonus_dietas = request.POST.get('bonus_dietas', None)
                data.kilometros = request.POST.get('kilometros', None)
                data.guardias = request.POST.get('guardias', None)
                data.tipo = request.POST.get('tipo', None)
                data.total = request.POST.get('total', None)
                data.incremento = request.POST.get('incremento', None)
                data.save()
        except:
            Sueldos(empleado=empleado,
                    anio=request.POST.get('anio', None),
                    retribucion_fija=request.POST.get('retribucion_fija', None),
                    varibale_individual=request.POST.get('varibale_individual', None),
                    varibale_empresa=request.POST.get('varibale_empresa', None),
                    beneficios_sociales=request.POST.get('beneficios_sociales', None),
                    bonus_dietas=request.POST.get('bonus_dietas', None),
                    kilometros=request.POST.get('kilometros', None),
                    guardias=request.POST.get('guardias', None),
                    total=request.POST.get('total', None),
                    tipo=request.POST.get('tipo', None),
                    incremento=request.POST.get('incremento', None)).save()
        finally:
            context['msg'] = "Sueldo añadido"
            context['form'] = new_form

    return render(request, 'empleados/sueldos.html', context)


@login_required(login_url='/accounts/login/')
@group_required('admin')
def ver_sueldos_propuestos(request):
    #aquellos sueldos del año siguiente que no pueden ver los empleados aun
    log_access(request)
    context = dict()
    context['title'] = 'Sueldos propuestos'
    if request.method == 'GET':
        logger.debug('%s - carga de información' % request.path)
        data = Sueldos_propuestos.objects.all()
        context['data'] = data

    return render(request, 'empleados/tabla_sueldos_propuestos.html', context)


@login_required(login_url='/accounts/login/')
@group_required('admin')
def modificar_sueldos_propuestos(request, id_empleado):
    #vista para modificar un sueldo propuesto ya guardado en bbdd
    log_access(request)
    context = dict()
    context['title'] = "Sueldos propuestos "
    context['subtitle'] = "Por favor, introduzca los datos numéricos sin puntos ni comas"
    context['id_empleado'] = id_empleado
    empleado = Empleados.objects.get(id_empleado=id_empleado)
    info = Sueldos_propuestos.objects.filter(empleado=id_empleado)
    if request.method == "POST":
        form = Sueldos_propuestos_modificarForm(request.POST, instance=info[0])
        if form.is_valid():
            form.save()
        return redirect('lista_sueldo_propuesto')
    else:
        form = Sueldos_propuestos_modificarForm(instance=info[0])
        context['form'] = form
        return render(request, 'empleados/modificar_propuesto.html', context)


@login_required(login_url='/accounts/login/')
@group_required('admin')
def aniadir_sueldo_propuesto(request):
    log_access(request)
    context = dict()
    new_form = Sueldos_propuestosForm
    context['title'] = "Sueldos propuestos"
    context['subtitle'] = "Por favor, introduzca los datos numéricos sin puntos ni comas"

    if request.method == 'GET':
        context['form'] = new_form
    elif request.method == 'POST':
        id_empleado = request.POST.get('empleado', None)
        empleado = Empleados.objects.get(id_empleado=id_empleado)
        try:
            data=Sueldos_propuestos.objects.get(empleado=empleado)
            if data:
                data.retribucion_fija=request.POST.get('retribucion_fija', None)
                data.varibale_individual=request.POST.get('varibale_individual', None)
                data.varibale_empresa=request.POST.get('varibale_empresa', None)
                data.beneficios_sociales=request.POST.get('beneficios_sociales', None)
                data.bonus_dietas=request.POST.get('bonus_dietas', None)
                data.kilometros=request.POST.get('kilometros', None)
                data.guardias=request.POST.get('guardias', None)
                data.tipo=request.POST.get('tipo', None)
                data.total=request.POST.get('total', None)
                data.incremento=request.POST.get('incremento', None)
                data.save()
        except:
            Sueldos_propuestos(empleado=empleado, retribucion_fija=request.POST.get('retribucion_fija', None),
                               varibale_individual=request.POST.get('varibale_individual', None),
                               varibale_empresa=request.POST.get('varibale_empresa', None),
                               beneficios_sociales=request.POST.get('beneficios_sociales', None),
                               bonus_dietas=request.POST.get('bonus_dietas', None),
                               kilometros=request.POST.get('kilometros', None),
                               guardias=request.POST.get('guardias', None),
                               tipo=request.POST.get('tipo', None),
                               total=request.POST.get('total', None),
                               incremento=request.POST.get('incremento', None)).save()
        finally:
            context['msg'] = "Sueldo propuesto añadido"
            context['form'] = new_form
    return render(request, 'empleados/sueldos_propuestos.html', context)


@login_required(login_url='/accounts/login/')
@group_required('admin')
def ver_sueldos(request):
    log_access(request)
    context = dict()
    context['title'] = 'Sueldos'
    now = datetime.now()
    if request.method == 'GET':
        logger.debug('%s - carga de información' % request.path)
        # Utilizamos Q para mas opciones de filtrado, en este caso distinto a.
        table_actual = SueldosTable(data=Sueldos.objects.all())
        context['table_actual'] = table_actual
    return render(request, 'empleados/tabla_sueldos.html', context)


@login_required(login_url='/accounts/login/')
@group_required('admin')
def reportes(request):
    log_access(request)
    context = dict()
    context['title'] = 'Reportes'

    logger.debug('%s - carga de pagina' % request.path)

    return render(request, 'reportes/reportes.html', context)


@login_required(login_url='/accounts/login/')
@group_required('admin')
def cargar_evaluaciones(request):
    log_access(request)
    context = dict()
    context['title'] = "Carga evaluacion"
    context['modelo'] = 3
    context['msg2'] = "ORDEN CSV: id_plantilla, año (YYYY), detalle, id_responsable, id_empleado, verificado(1=si,0=no)," \
                      " fecha_verificado(YYYY-MM-DD), observaciones, externo. Los tres últimas campos no son obligatorios, pueden ir vacíos."

    if request.method == 'GET':
        logger.debug('/subida_archivo_evaluacion/ - carga inicial')
    elif request.method == 'POST':
        try:
            if not 'file' in request.FILES:
                context['msg'] = "ERROR: No ha seleccionado archivo."
            elif not str(request.FILES['file']).lower().endswith('.csv'):
                context['msg'] = "ERROR: No es un archivo CSV."
            else:
                if processFileEvaluacion(csv_file=request.FILES['file']):
                    context['msg'] = "Evaluacion creada"
                else:
                    context['msg'] = "Error al subir evaluaciones, no es un fichero csv"
        except AttributeError:
            context['msg'] = 'ERROR: Formato incorrecto'
        except StopIteration:
            context['msg'] = 'ERROR: El fichero está vacío'
        except Exception as exception:
            context['msg'] = 'ERROR: ' + type(exception).__name__ + ': ' + str(exception.__str__())

    return render(request, 'evaluaciones/carga_evaluaciones.html', context)


@login_required(login_url='/accounts/login/')
@group_required('admin')
def cargar_preguntas(request):
    log_access(request)
    context = dict()
    context['title'] = "Carga preguntas"
    context['modelo'] = 4
    context['msg2'] = "ORDEN CSV: id_plantilla, id_seccion, pregunta"

    if request.method == 'GET':
        logger.debug('/subida_archivo_preguntas/ - carga inicial')
    elif request.method == 'POST':
        try:
            if not 'file' in request.FILES:
                context['msg'] = "ERROR: No ha seleccionado archivo."
            elif not str(request.FILES['file']).lower().endswith('.csv'):
                context['msg'] = "ERROR: No es un archivo CSV."
            else:
                if processFilePreguntas(csv_file=request.FILES['file']):
                    context['msg'] = "Preguntas creadas"
                else:
                    context['msg'] = "Error al subir preguntas, no es un fichero csv"
        except AttributeError:
            context['msg'] = 'ERROR: Formato incorrecto'
        except StopIteration:
            context['msg'] = 'ERROR: El fichero está vacío'
        except Exception as exception:
            context['msg'] = 'ERROR: ' + type(exception).__name__ + ': ' + str(exception.__str__())

    return render(request, 'preguntas/cargar_preguntas.html', context)


@login_required(login_url='/accounts/login/')
@group_required('admin')
def cargar_respuestas(request):
    log_access(request)
    context = dict()
    context['title'] = "Carga respuestas"
    context['modelo'] = 5
    context['msg2'] = "ORDEN CSV: id_evaluacion, id_pregunta, respuestas"

    if request.method == 'GET':
        logger.debug('/subida_archivo_respuestas/ - carga inicial')
    elif request.method == 'POST':
        try:
            if not 'file' in request.FILES:
                context['msg'] = "ERROR: No ha seleccionado archivo."
            elif not str(request.FILES['file']).lower().endswith('.csv'):
                context['msg'] = "ERROR: No es un archivo CSV."
            else:
                if processFileRespuestas(csv_file=request.FILES['file']):
                    context['msg'] = "Respuestas creadas"
                else:
                    context['msg'] = "Error al subir respuestas, no es un fichero csv"
        except AttributeError:
            context['msg'] = 'ERROR: Formato incorrecto'
        except StopIteration:
            context['msg'] = 'ERROR: El fichero está vacío'
        except Exception as exception:
            context['msg'] = 'ERROR: ' + type(exception).__name__ + ': ' + str(exception.__str__())

    return render(request, 'preguntas/cargar_respuestas.html', context)


@login_required(login_url='/accounts/login/')
@group_required('admin')
def organigrama(request):
    context = dict()
    context['title'] = 'Organigrama'
    ceo = Empleados.objects.get(responsable__exact = F('id_empleado'))
    #ceo = Empleados.objects.get(responsable=None)
    #print('***********************',ceo)
    jefes_depart = Empleados.objects.filter(resp_depart=True)
    organigrama = {}
    empleados = [ceo]
    #print('***********************',empleados)

    lista_datos = list()
    organigrama_depart = dict()

    # Organigrama general
    #print('__________*************Entra a formar el organigrama general')
    organigrama_depart['ORGANIGRAMA GENERAL'] = crearListaOrganigrama(insert_new_level(empleados, organigrama),lista_datos)

                                                                    
    # Organigramas departamentos
   #print('__________***********Entra a formar el organigrama departamentos')
    lista_depart = list()

    for jefe in jefes_depart:
        if jefe != ceo:
            jefe_depart = [jefe]
            organigrama = {}
            organigrama_depart['ORGANIGRAMA ÁREA ' + jefe.area_empleado] = crearListaOrganigrama(
                insert_new_level(jefe_depart, organigrama),
                lista_depart)
            lista_depart = list()
    context['org_depart'] = organigrama_depart

    return render(request, 'organigrama/organigrama.html', context)


# Descarga pdf de evaluacion
@login_required(login_url='/accounts/login/')
@group_required('admin', 'responsable')
def download_pdf(request, nombre):
    log_access(request)
    if not nombre:
        return render(request, '404.html')

    file_path = os.path.join(MEDIA_ROOT, nombre + '.pdf')
    logger.debug('/download_pdf/ %s' % file_path)
    if os.path.exists(file_path):
        usuario = request.user.username
        actividad = Actividades(usuario=usuario, accion='DESCARGA PDF',
                                informacion=nombre)
        actividad.save()
        with open(file_path, 'rb') as fh:
            response = HttpResponse(fh.read(), content_type="application/pdf")
            response['Content-Disposition'] = 'inline; filename=' + os.path.basename(file_path)
            return response
    else:
        logger.error('%s not exists' % file_path)
        return render(request, '404.html')


# Descarga pdf de empleados a cargo del responsable
@login_required(login_url='/accounts/login/')
@group_required('admin', 'responsable')
def download_pdf_rp(request, id):
    log_access(request)
    if not id:
        return render(request, '404.html')

    comprobacion = comprobar_identificador(request, id)

    if comprobacion:
        file_path = os.path.join(MEDIA_ROOT, id + '.pdf')
        logger.debug('/download_pdf/ %s' % file_path)
        if os.path.exists(file_path):
            usuario = request.user.username
            actividad = Actividades(usuario=usuario, accion='DESCARGAR PDF RP',
                                    informacion='')
            actividad.save()
            with open(file_path, 'rb') as fh:
                response = HttpResponse(fh.read(), content_type="application/pdf")
                response['Content-Disposition'] = 'inline; filename=' + os.path.basename(file_path)
                return response
        else:
            logger.error('%s not exists' % file_path)
            return render(request, '404.html')
    else:
        return render(request, '404.html')


# Descarga pdf de empleados a cargo del responsable de proyecto
@login_required(login_url='/accounts/login/')
@group_required('admin', 'responsable')
def download_pdf_ep(request, id, ep):
    log_access(request)
    if not id:
        return render(request, '404.html')

    comprobacion = False

    if ep:
        comprobacion = comprobar_identificador_ep(request, id)

    if comprobacion:
        file_path = os.path.join(MEDIA_ROOT, id + '_ep.pdf')
        logger.debug('/download_pdf_ep/ %s' % file_path)
        if os.path.exists(file_path):
            usuario = request.user.username
            actividad = Actividades(usuario=usuario, accion='DESCARGA PDF EP',
                                    informacion='')
            actividad.save()
            with open(file_path, 'rb') as fh:
                response = HttpResponse(fh.read(), content_type="application/pdf")
                response['Content-Disposition'] = 'inline; filename=' + os.path.basename(file_path)
                return response
        else:
            logger.error('%s not exists' % file_path)
            return render(request, '404.html')
    else:
        return render(request, '404.html')


@login_required(login_url='/accounts/login/')
@group_required('admin')
def ver_departamentos(request):
    context = dict()
    mis_empleados_datos = dict()
    anio_seleccionados = list()
    empleado_post = list()
    sueldos_anios = list()
    new_form = ListaDepartamentoForm
    tipoSueldo_seleccionados = ""
    context['title'] = "Departamentos"
    context['form'] = new_form
    fecha_actual = datetime.now()
    anio_seleccionados.append(fecha_actual.year)

    global departamento
    if request.method == 'GET':
        context['get'] = True
    elif request.method == 'POST':

        if 'buscarDept' in request.POST:
            departamento_post = ref_areas.objects.get(id_area=request.POST.get('departamento', None))
            departamento = departamento_post.descripcion
            empleados_departamento = Empleados.objects.filter(area_empleado=departamento)
            for empleado in empleados_departamento:
                empleado_post.append(empleado)

        else:
            empleados_departamento = Empleados.objects.filter(area_empleado=departamento)
            if 'empleado' in request.POST:
                empleados_POST = request.POST.getlist('empleado')
                for empleado in empleados_POST:
                    empleado_post.append(Empleados.objects.get(email=empleado))
            else:
                for empleado in empleados_departamento:
                    empleado_post.append(empleado)

            if 'anio' in request.POST:
                anio_seleccionados.clear()
                anio_seleccionados = request.POST.getlist('anio')
            if 'tipoSueldo' in request.POST:
                tipoSueldo_seleccionados = request.POST.getlist('tipoSueldo')

        for empleado in empleados_departamento:
            enlace = "/empleado/detalle/" + str(empleado.id_empleado)
            mis_empleados_datos[empleado] = enlace

        # Sueldos
        anios_sueldos = Sueldos.objects.filter(empleado__in=empleados_departamento)
        sueldos = Sueldos.objects.filter(empleado__in=empleado_post)

        for sueldo in anios_sueldos:
            if sueldo.anio not in sueldos_anios:
                sueldos_anios.append(sueldo.anio)

        tablaSueldo_propuesto = SueldosTable(
            data=Sueldos.objects.filter(empleado__in=empleado_post, anio__in=anio_seleccionados))
        tablaSueldo_propuesto.paginate(page=request.GET.get('page', 1), per_page=PER_PAGE)

        context['misEmpleados'] = mis_empleados_datos
        RequestConfig(request).configure(tablaSueldo_propuesto)  # to sort
        context['departamento'] = departamento
        context['misEmpleados'] = mis_empleados_datos
        context['tablaSueldo_propuesto'] = tablaSueldo_propuesto
        context['seleccionSueldo'] = VistaSueldosForm(empleados_departamento, sueldos, sueldos_anios)
        context['grafico'] = chartResponsable(empleado_post, tipoSueldo_seleccionados, anio_seleccionados)

    return render(request, 'departamentos/detalle_departamentos.html', context)


def log_access(request):
    try:
        logger.debug('ACCESS|%s|%s' % (request.user.username, request.path))
    except:
        logger.error('ACCESS Request sin informacion')


def guardar_foto(request, identificador):
    #no se guarda la foto, no encuentro el error
    try:
        if not 'photo' in request.FILES:
            logger.debug('no foto subida')
            return False
        elif not str(request.FILES['photo']).lower().endswith('.jpeg'):
            return False
        else:
            if os.path.isfile(base.MEDIA_ROOT + '/photos/' + identificador + '.jpeg'):
                os.remove(base.MEDIA_ROOT + '/photos/' + identificador + '.jpeg')
            save_path = os.path.join(base.MEDIA_ROOT, 'photos', identificador + '.jpeg')
            path = default_storage.save(save_path, request.FILES['photo'])
            default_storage.path(path)
            usuario = request.user.username
            actividad = Actividades(usuario=usuario, accion='CARGA FOTO PERFIL',
                                    informacion=identificador)
            actividad.save()
            return True
    except Exception as exception:
        print('ERROR: ' + type(exception).__name__ + ': ' + str(exception.__str__()))
        return False


@login_required(login_url='/accounts/login/')
@group_required('admin', 'empleado')
def responsable_evaluaciones(request, id_empleado):
    if comprobar_usuario(request, id_empleado):
        if request.method == 'GET':
            fecha = datetime.now()
            context = dict()
            context['id_empleado'] = id_empleado
            #context['title'] = 'active_evaluaciones'
            empleado = Empleados.objects.get(id_empleado=id_empleado)
            context['nombre'] = empleado.nombre + " " + empleado.apellidos
            context['name_menu'] = empleado.nombre
            context['identificador'] = empleado.identificador
            context['anio_actual'] = fecha.year

            # Comprobacion de permisos de visualizacion
            if empleado.evaluador or empleado.resp_depart or empleado.email == 'admin':
                context['permiso'] = True
                context['menu'] = True
            if empleado.ver_sueldos:
                context['ver_sueldo'] = True

            # Buscar evaluaciones del empleado
            plantilla_eval = PlantillaEvaluacion.objects.filter(autoevaluacion=False, de_responsable=False)
            #return HttpResponse(plantilla_eval)
            
            if empleado.resp_depart:
                context['resp_depart'] = True
                evaluaciones_responsable = Evaluacion.objects.filter(
                    empleado__in=Empleados.objects.filter(area_empleado=empleado.area_empleado), anio=fecha.year,id_plantilla__in=plantilla_eval).order_by('empleado__apellidos')

                # EVALUACIONES DEL DEPARTAMENTO
                tableEvalDepart = EvaluacionRespTable(
                    data=evaluaciones_responsable)
                tableEvalDepart.exclude = ('id_evaluacion', 'id_plantilla', 'anio','fecha_verificado', 'observaciones')
                #tableEvalDepart.paginate(page=request.GET.get('page', 1), per_page=PER_PAGE)
                context['tableEvalDepart'] = tableEvalDepart
                RequestConfig(request,paginate={"per_page": 5}).configure(tableEvalDepart)


            evaluaciones_empleado = Evaluacion.objects.filter(responsable=empleado, anio=fecha.year,
                                                              id_plantilla__in=plantilla_eval).order_by('empleado__apellidos')
            #return HttpResponse(evaluaciones_empleado)
            

            context['progreso'] = calcularProgreso(empleado)
            context['n_evaluaciones'] = len(evaluaciones_empleado)

            # EVALUACIONES ACTUALES
            tablaEvaluaciones = EvaluacionRespTable(
                data=evaluaciones_empleado)
            tablaEvaluaciones.exclude = (
                'id_evaluacion', 'id_plantilla','anio','fecha_verificado', 'responsable', 'observaciones')
            #tablaEvaluaciones.paginate(page=request.GET.get('page', 1), per_page=PER_PAGE)
            context['tableEvalActual'] = tablaEvaluaciones
            #return HttpResponse (tablaEvaluaciones)
            RequestConfig(request,paginate={"per_page": 10}).configure(tablaEvaluaciones)

            #AUTOEVALUACIONES de empleados a su cargo ordenadas por apellido de empleado
            plantilla_auteval = PlantillaEvaluacion.objects.filter(autoevaluacion=True, de_responsable=False)
            
            autoevaluaciones = Evaluacion.objects.filter(responsable=empleado, anio=fecha.year,
                                                              id_plantilla__in=plantilla_auteval).order_by('empleado__apellidos')
            #return HttpResponse(autoevaluaciones) 
            
            tablaAutoevaluaciones = EvaluacionRespTable(data=autoevaluaciones)

            tablaAutoevaluaciones.exclude = ('id_evaluacion', 'id_plantilla','anio', 'fecha_verificado','responsable', 'observaciones')
            context['tablaAutoeval'] = tablaAutoevaluaciones
            RequestConfig(request,paginate={"per_page": 10}).configure(tablaAutoevaluaciones)
            

            # EVALUACIONES PASADAS
            #tablaEvaluacionesPasadas = EvaluacionRespTable(
                #data=Evaluacion.objects.filter(~Q(anio=fecha.year), responsable=empleado,
                                               #id_plantilla__in=plantilla_eval).order_by('empleado__nombre'))

            plantilla_anteriores= Evaluacion.objects.filter(~Q(anio=fecha.year), responsable=empleado).order_by('anio','empleado__apellidos').reverse()
            tablaEvaluacionesPasadas = EvaluacionRespTable(data=plantilla_anteriores)
            tablaEvaluacionesPasadas.exclude = ('id_evaluacion', 'id_plantilla', 'fecha_verificado', 'responsable', 'observaciones')
            #tablaEvaluacionesPasadas.paginate(page=request.GET.get('page', 1), per_page=PER_PAGE)
            context['tableEvalPasada'] = tablaEvaluacionesPasadas
            RequestConfig(request,paginate={"per_page": 5}).configure(tablaEvaluacionesPasadas)
            #return HttpResponse(plantilla_anteriores) 
            usuario = request.user.username
            actividad = Actividades(usuario=usuario, accion='CONSULTA DEL PROGRESO EVALUACIONES POR EMPLEADOS',
                                    informacion='')
            actividad.save()
            return render(request, 'empleado/resp_evaluaciones.html', context)
    else:
        return render(request, '404.html')


@login_required(login_url='/accounts/login/')
@group_required('admin', 'responsable')
def responsable_sueldos(request, id_empleado):
    if comprobar_usuario(request, id_empleado):
        context = dict()
        misEmpleadosEval = list()
        anio_seleccionados = list()
        tipoSueldo_seleccionados = ""
        fecha_actual = datetime.now()
        empleado = Empleados.objects.get(id_empleado=id_empleado)
        if empleado.resp_depart:
            misEmpleados = Empleados.objects.filter(area_empleado=empleado.area_empleado, eliminado=False)
        else:
            misEmpleados = Empleados.objects.filter(responsable=empleado, eliminado=False)
        context['title'] = 'Responsable datos'
        context["empDetalle"] = True

        if empleado.evaluador or empleado.resp_depart or empleado.email == 'admin':
            context['permiso'] = True
            context['menu'] = True
        if empleado.ver_sueldos:
            context['ver_sueldo'] = True

        if empleado.evaluador:
            context['nombre'] = empleado.nombre + ' ' + empleado.apellidos
            context['name_menu'] = empleado.nombre
            context['id_empleado'] = empleado.id_empleado
            context['identificador'] = empleado.identificador
            context['sueldo_max'] = calcularSMax(empleado)
            context['sueldo_medio'] = calcularSMedio(empleado)
            context['sueldo_min'] = calcularSMin(empleado)

            list_ids_plantilla = PlantillaEvaluacion.objects.filter().values_list(
                'id_plantilla',
                flat=True)
            if empleado.ver_sueldos:
                context['ver_sueldos'] = True

            # Buscar empleados
            if misEmpleados:
                for empleado in misEmpleados:
                    misEmpleadosEval.append(empleado)
            else:
                # En este caso, misEmpleado contiene evaluaciones por eso el if empleado.empleado de despues.
                misEvaluaciones = Evaluacion.objects.filter(responsable=empleado.id_empleado,
                                                            id_plantilla__in=list_ids_plantilla)
                for empleado in misEvaluaciones:
                    misEmpleadosEval.append(empleado.empleado)

        # Sueldos
        sueldos = Sueldos.objects.filter(empleado__in=misEmpleadosEval)

        sueldos_anios = list()
        for sueldo in sueldos:
            if sueldo.anio not in sueldos_anios:
                sueldos_anios.append(sueldo.anio)
        context['seleccionSueldo'] = VistaSueldosForm(misEmpleados, sueldos, sueldos_anios)
        if request.method == 'GET':
            tablaSueldo = SueldosTable(
                data=Sueldos.objects.filter(empleado__in=misEmpleadosEval, anio=fecha_actual.year, tipo="Actual"))
            tablaSueldo.paginate(page=request.GET.get('page', 1), per_page=PER_PAGE)
            context['tablaSueldo'] = tablaSueldo
        elif request.method == 'POST':
            misEmpleadosEval = list()
            if 'empleado' in request.POST:
                misEmpleadosEval.clear()
                empleados_POST = request.POST.getlist('empleado')
                for empleado in empleados_POST:
                    misEmpleadosEval.append(Empleados.objects.get(email=empleado))
            else:
                empleados_departamento = Empleados.objects.filter(area_empleado=empleado.area_empleado)
                for empleado in empleados_departamento:
                    misEmpleadosEval.append(empleado)
            if 'anio' in request.POST:
                anio_seleccionados = request.POST.getlist('anio')
            else:
                anio_seleccionados = list()
                anio_seleccionados.append(fecha_actual.year)
            if 'tipoSueldo' in request.POST:
                tipoSueldo_seleccionados = request.POST.getlist('tipoSueldo')

            tablaSueldo = SueldosTable(
                data=Sueldos.objects.filter(empleado__in=misEmpleadosEval, anio__in=anio_seleccionados))
            tablaSueldo.paginate(page=request.GET.get('page', 1), per_page=PER_PAGE)
            context['tablaSueldo'] = tablaSueldo
        context['grafico'] = chartResponsable(misEmpleadosEval, tipoSueldo_seleccionados, anio_seleccionados)
        usuario = request.user.username
        actividad = Actividades(usuario=usuario, accion='CONSULTA SUELDOS ACTUALES DE EMPLEADOS',
                                informacion='')
        actividad.save()
        return render(request, 'empleado/resp_sueldos.html', context)
    else:
        return render(request, '404.html')

@login_required(login_url='/accounts/login/')
@group_required('admin', 'responsable', 'ver_sueldos')
def responsable_sueldos_propuestos(request, id_empleado):
    if comprobar_usuario(request, id_empleado):
        context = dict()
        misEmpleadosEval = list()
        tipoSueldo_seleccionados = ""
        empleado = Empleados.objects.get(id_empleado=id_empleado)
        if empleado.resp_depart:
            misEmpleados = Empleados.objects.filter(area_empleado=empleado.area_empleado, eliminado=False)
        else:
            misEmpleados = Empleados.objects.filter(responsable=empleado, eliminado=False)
        context['title'] = 'Responsable datos'
        context["empDetalle"] = True

        if empleado.evaluador or empleado.resp_depart or empleado.email == 'admin':
            context['permiso'] = True
            context['menu'] = True
        if empleado.ver_sueldos:
            context['ver_sueldo'] = True

        if empleado.evaluador:
            context['nombre'] = empleado.nombre + ' ' + empleado.apellidos
            context['name_menu'] = empleado.nombre
            context['id_empleado'] = empleado.id_empleado
            context['identificador'] = empleado.identificador


            list_ids_plantilla = PlantillaEvaluacion.objects.filter().values_list('id_plantilla',flat=True)
            if empleado.ver_sueldos:
                context['ver_sueldos'] = True

            # Buscar empleados
            if misEmpleados:
                for empleado in misEmpleados:
                    misEmpleadosEval.append(empleado)
            else:
                # En este caso, misEmpleado contiene evaluaciones por eso el if empleado.empleado de despues.
                misEvaluaciones = Evaluacion.objects.filter(responsable=empleado.id_empleado,
                                                            id_plantilla__in=list_ids_plantilla)
                for empleado in misEvaluaciones:
                    misEmpleadosEval.append(empleado.empleado)

        # Sueldos
        sueldos = Sueldos.objects.filter(empleado__in=misEmpleadosEval)
        context['seleccionSueldo'] = VistaSueldos_propuestosForm(misEmpleados,sueldos)

        if request.method == 'GET':
            tablaSueldo_propuesto = Sueldos_propuestosTable(
                data=(Sueldos_propuestos.objects.filter(empleado__in=misEmpleadosEval)))
            tablaSueldo_propuesto.paginate(page=request.GET.get('page', 1), per_page=PER_PAGE)
            context['tablaSueldo_propuesto'] = (tablaSueldo_propuesto)
        elif request.method == 'POST':
            misEmpleadosEval = list()
            if 'empleado' in request.POST:
                misEmpleadosEval.clear()
                empleados_POST = request.POST.getlist('empleado')
                for empleado in empleados_POST:
                    misEmpleadosEval.append(Empleados.objects.get(email=empleado))
            else:
                empleados_departamento = Empleados.objects.filter(area_empleado=empleado.area_empleado)
                for empleado in empleados_departamento:
                    misEmpleadosEval.append(empleado)
            if 'tipoSueldo' in request.POST:
                tipoSueldo_seleccionados = request.POST.getlist('tipoSueldo')


            # Tabla sueldos propuesto
            tablaSueldo_propuesto = Sueldos_propuestosTable(
                data=Sueldos_propuestos.objects.filter(empleado__in=misEmpleadosEval))
            tablaSueldo_propuesto.paginate(page=request.GET.get('page', 1), per_page=PER_PAGE)
            context['tablaSueldo_propuesto'] = tablaSueldo_propuesto
        context['grafico'] = chartResponsable_propuesto(misEmpleadosEval, tipoSueldo_seleccionados)
        usuario = request.user.username
        actividad = Actividades(usuario=usuario, accion='CONSULTA SUELDOS PROPUESTOS DE EMPLEADOS',
                                informacion='')
        actividad.save()
        return render(request, 'empleado/resp_sueldos_propuestos.html', context)
    else:
        return render(request, '404.html')


@login_required(login_url='/accounts/login/')
@group_required('admin', 'empleado')
def empleado_perfil(request, id_empleado):
    if comprobar_usuario(request, id_empleado):
        context = dict()
        fecha = datetime.now()
        page = request.GET.get('page', 1)
        if request.user.username == 'admin':
            empleado = Empleados.objects.get(id_empleado=id_empleado)
        else:
            empleado = Empleados.objects.get(email=request.user.username)


        # Datos empleado
        context['id_empleado'] = empleado.id_empleado
        context['nombre'] = empleado.nombre + " " + empleado.apellidos
        context['name_menu'] = empleado.nombre
        if empleado.responsable:
            context['nombre_responsable'] = empleado.responsable.nombre + " " + empleado.responsable.apellidos
            context['correo_responsable'] = empleado.responsable.email
            context['id_responsable'] = empleado.responsable.id_empleado
        # context['email'] = empleado.email
        context['identificador'] = empleado.id_empleado
        context['depart'] = empleado.area_empleado
        context['antiguedad'] = empleado.f_antiguedad
        context['rol'] = empleado.rol

        if empleado.evaluador or empleado.resp_depart or empleado.email == 'admin':
            context['permiso'] = True
            context['menu'] = True
        if empleado.ver_sueldos:
            context['ver_sueldo'] = True

        # Formulario conocimientos
        new_form = EmpleadoNuevoConocimientoForm(initial={"id_empleado": empleado.id_empleado})

        new_form.fields['otroConocimiento'].widget = forms.TextInput(
            attrs={'placeholder': 'Introduce aquí el conocimiento'})
        if request.method == 'GET':
            # Autoevaluaciones
            tablaAutoevaluaciones = EvaluacionesPerfil(
                data=Evaluacion.objects.filter(empleado=empleado, anio=fecha.year,
                                               id_plantilla__in=PlantillaEvaluacion.objects.filter(
                                                   autoevaluacion=True, de_responsable=False)))
            tablaAutoevaluaciones.exclude = (
                'id_evaluacion', 'id_plantilla', 'fecha_verificado', 'responsable', 'actualizado', 'observaciones')
            context['tablaAutoevaluaciones'] = tablaAutoevaluaciones

            tablaAutovaluacionesPasadas = EvaluacionesPerfil(
                data=Evaluacion.objects.filter(~Q(anio=fecha.year), empleado=empleado,
                                               id_plantilla__in=PlantillaEvaluacion.objects.filter(
                                                   autoevaluacion=True)))
            tablaAutovaluacionesPasadas.exclude = (
                'id_evaluacion', 'id_plantilla', 'fecha_verificado', 'responsable', 'actualizado', 'observaciones')
            context['tablaAutovaluacionesPasadas'] = tablaAutovaluacionesPasadas

            # Evaluaciones
            tablaEvaluaciones = EvaluacionesPerfil(
                data=Evaluacion.objects.filter(empleado=empleado, anio=fecha.year,
                                               id_plantilla__in=PlantillaEvaluacion.objects.filter(
                                                   autoevaluacion=False, de_responsable=False)))
            tablaEvaluaciones.exclude = (
                'id_evaluacion', 'id_plantilla', 'fecha_verificado', 'responsable', 'actualizado', 'observaciones')
            context['tablaEvaluaciones'] = tablaEvaluaciones

            tablaEvaluacionesPasadas = EvaluacionesPerfil(
                data=Evaluacion.objects.filter(~Q(anio=fecha.year), empleado=empleado,
                                               id_plantilla__in=PlantillaEvaluacion.objects.filter(
                                                   autoevaluacion=False, de_responsable=False)))
            tablaEvaluacionesPasadas.exclude = (
                'id_evaluacion', 'id_plantilla', 'fecha_verificado', 'responsable', 'actualizado', 'observaciones')
            context['tablaEvaluacionesPasadas'] = tablaEvaluacionesPasadas

            # Evaluacion al responsable
            try:
                evaluacion_responsable = Evaluacion.objects.get(anio=fecha.year, empleado=empleado,
                                                                id_plantilla__in=PlantillaEvaluacion.objects.filter(
                                                                    autoevaluacion=False, de_responsable=True))
                context['evaluacion_al_responsable'] = '/evaluacion/preguntas/lista/' + str(
                    evaluacion_responsable.id_evaluacion) + '/' + str(
                    empleado.id_empleado)
            except:
                evaluacion_responsable = ""
                context['evaluacion_al_responsable'] = evaluacion_responsable

            # Grafico sueldos.
            context['grafico'] = chartEmpleado(int(empleado.id_empleado))

            if empleado.evaluador:
                # Grafico notas
                context['graficoNotas'] = chart_eval_responsable(empleado)

            # Conocimientos
            conocimientos = Conocimientos.objects.filter(id_empleado=empleado.id_empleado)
            paginator = Paginator(conocimientos, 5)
            try:
                conocimiento_paginate = paginator.page(page)
            except PageNotAnInteger:
                conocimiento_paginate = paginator.page(1)
            except EmptyPage:
                conocimiento_paginate = paginator.page(paginator.num_pages)

            context['conocimiento_paginate'] = conocimiento_paginate
            conocimiento_list = list()
            for conocimiento in conocimientos:
                conocimiento_list.append(conocimiento.descripcion)
            context['conocimientos'] = conocimiento_list

            context['form'] = new_form
            new_form.fields['id_empleado'].widget = forms.HiddenInput()
            usuario = request.user.username
            actividad = Actividades(usuario=usuario, accion='VISTA PERFIL',
                                    informacion='')
            actividad.save()
            return render(request, 'empleado/emp_perfil.html', context)
        elif 'Crear' in request.POST:
            conocimiento = ref_Conocimiento.objects.get(
                id_conocimiento=request.POST.get('descripcion', None))
            if conocimiento.descripcion == "Otro":
                # Comprobar que el conocimiento no existe
                buscarConocimiento = Conocimientos.objects.filter(
                    descripcion=request.POST.get('otroConocimiento'),
                    id_empleado=empleado)
                nuevoConocimiento = request.POST.get('otroConocimiento')
            else:
                # Comprobar que el conocimiento no existe
                buscarConocimiento = Conocimientos.objects.filter(
                    descripcion=conocimiento.descripcion,
                    id_empleado=empleado)
                nuevoConocimiento = conocimiento.descripcion
            if buscarConocimiento:
                Conocimientos.objects.filter(
                    descripcion=nuevoConocimiento,
                    id_empleado=empleado,
                ).update(
                    nivel=request.POST.get('nivel', None),
                    certificado=get_checkbox_value(request.POST.get('certificado', None))
                )
                usuario = request.user.username
                actividad = Actividades(usuario=usuario, accion='ACTUALIZAR CONOCIMIENTO',
                                        informacion=nuevoConocimiento)
                actividad.save()
            else:
                Conocimientos(descripcion=nuevoConocimiento,
                              nivel=request.POST.get('nivel', None),
                              id_empleado=empleado,
                              certificado=get_checkbox_value(request.POST.get('certificado', None))).save()
                usuario = request.user.username
                actividad = Actividades(usuario=usuario, accion='NUEVO CONOCIMIENTO',
                                        informacion=nuevoConocimiento)
                actividad.save()
            context['form'] = new_form
            new_form.fields['id_empleado'].widget = forms.HiddenInput()

        elif 'Actualizar' in request.POST:
            conocimiento = ref_Conocimiento.objects.get(
                id_conocimiento=request.POST.get('descripcion', None))
            if conocimiento.descripcion == "Otro":

                Conocimientos.objects.filter(
                    descripcion=request.POST.get('otroConocimiento', None),
                    id_empleado=empleado,
                ).update(
                    nivel=request.POST.get('nivel', None),
                    certificado=get_checkbox_value(request.POST.get('certificado', None))
                )
                usuario = request.user.username
                actividad = Actividades(usuario=usuario, accion='ACTUALIZAR CONOCIMIENTO',
                                        informacion=conocimiento)
                actividad.save()
            else:
                Conocimientos.objects.filter(
                    descripcion=ref_Conocimiento.objects.get(
                        descripcion=conocimiento),
                    id_empleado=empleado,
                ).update(
                    nivel=request.POST.get('nivel', None),
                    certificado=get_checkbox_value(request.POST.get('certificado', None))
                )
                usuario = request.user.username
                actividad = Actividades(usuario=usuario, accion='ACTUALIZAR CONOCIMIENTO',
                                        informacion=conocimiento)
                actividad.save()

            context['form'] = new_form
            new_form.fields['id_empleado'].widget = forms.HiddenInput()
        elif 'Borrar' in request.POST:
            usuario = request.user.username
            conocimiento = Conocimientos.objects.get(id_conocimiento=request.POST['id_conocimiento'])
            actividad = Actividades(usuario=usuario, accion='BORRAR CONOCIMIENTO',
                                    informacion=conocimiento)
            actividad.save()
            Conocimientos.objects.filter(
                id_conocimiento=request.POST['id_conocimiento']).delete()
            pass

        return redirect('empleadoPerfil', id_empleado=id_empleado)
    else:
        return render(request, '404.html')


# TODO: BORRAR AL ACABAR
@login_required(login_url='/accounts/login/')
@group_required('admin', 'empleado')
def evaluacion_respuestas2(request, id_evaluacion, id_empleado, siguiente_pregunta):
    log_access(request)
    if comprobar_responsable(request, id_evaluacion):
        context = dict()
        int_siguiente_pregunta = int(siguiente_pregunta)

        # Datos de la evaluacion
        int_idempleado = int(id_empleado)
        int_idevaluacion = int(id_evaluacion)

        # Datos empleado logueado
        if request.user.username == 'admin':
            empleado_log = Empleados.objects.get(id_empleado=id_empleado)
        else:
            empleado_log = Empleados.objects.get(email=request.user.username)

        context['id_empleado'] = empleado_log.id_empleado
        context['nombre'] = empleado_log.nombre + ' ' + empleado_log.apellidos
        context['name_menu'] = empleado_log.nombre
        context['identificador'] = empleado_log.identificador
        if empleado_log.evaluador or empleado_log.resp_depart or empleado_log.email == 'admin':
            context['permiso'] = True
            context['menu'] = True
        if empleado_log.ver_sueldos:
            context['ver_sueldo'] = True
        # Datos empleado
        empleado = Empleados.objects.get(id_empleado=int_idempleado)

        context['id_empleado_find'] = empleado.id_empleado
        context['nombre_find'] = empleado.nombre + ' ' + empleado.apellidos
        context['email_find'] = empleado.email
        context['depart_find'] = empleado.area_empleado
        context['identificador_find'] = empleado.identificador
        context['antiguedad_find'] = empleado.f_antiguedad
        context['grafico'] = chartEmpleado(id_empleado)

        # Datos de la evaluacion
        evaluacion = Evaluacion.objects.get(id_evaluacion=int_idevaluacion)
        if evaluacion.id_plantilla.autoevaluacion:
            context['titulo'] = "AUTO EVALUACION"
            plantilla = PlantillaEvaluacion.objects.get(id_plantilla=evaluacion.id_plantilla.id_plantilla)
        else:
            context['titulo'] = "EVALUACION DE DESEMPEÑO"
            plantilla = PlantillaEvaluacion.objects.get(id_plantilla=evaluacion.id_plantilla.id_plantilla,
                                                        autoevaluacion=False)

        secciones = Secciones.objects.all()
        id_preguntas = list()
        lista_preguntas = list()

        context['id_responsable'] = evaluacion.responsable.id_empleado
        context['id_evaluacion'] = evaluacion.id_evaluacion

        if evaluacion.responsable != empleado.responsable:
            context['respEvaluacion'] = True

        for seccion in secciones:
            lista_preguntas.append(
                Preguntas.objects.filter(id_plantilla=plantilla.id_plantilla, id_seccion=seccion.id_seccion))

        for pregunta_seccion in lista_preguntas:
            for pregunta in pregunta_seccion:
                id_preguntas.append(pregunta.id_pregunta)  # Lista con los id de las preguntas de la evaluacion

        try:
            observaciones = evaluacion.observaciones
            if observaciones == None:
                context['txtobservaciones'] = ""
            else:
                context['txtobservaciones'] = observaciones
        except:
            context['txtobservaciones'] = ""

        if request.method == 'GET':

            pregunta_mostrar = Preguntas.objects.get(id_pregunta=int_siguiente_pregunta)
            # Buscar si ya ha introducido anteriormente una respuesta
            try:
                respuesta = Respuestas.objects.get(id_pregunta=pregunta_mostrar.id_pregunta,
                                                   id_evaluacion=evaluacion.id_evaluacion)
                respuesta_text = respuesta.respuesta
            except:
                respuesta_text = ""

            enunciado = pregunta_mostrar.pregunta  # Coger el enunciado de la pregunta pasada
            seccion = pregunta_mostrar.id_seccion.descripcion  # Coger la seccion de la pregunta
            new_form = RespuestaForm(
                initial={"respuesta": respuesta_text, "id_pregunta": pregunta_mostrar.id_pregunta,
                         "id_evaluacion": id_evaluacion})
            new_form.fields['id_pregunta'].widget = forms.HiddenInput()
            new_form.fields['id_evaluacion'].widget = forms.HiddenInput()
            context['form'] = new_form
            context['enunciado'] = enunciado
            context['seccion'] = seccion
            context['siguiente_pregunta'] = pregunta_mostrar.id_pregunta
            return render(request, 'evaluaciones/evaluacion_respuesta.html', context)
        elif request.method == 'POST':
            context['tipo'] = "responsable"
            Evaluacion.objects.filter(id_evaluacion=id_evaluacion).update(
                observaciones=request.POST.get('obs', None))
            # Guardar o actualizar respuesta
            try:
                Respuestas(respuesta=request.POST['respuesta'],
                           id_pregunta=Preguntas.objects.get(id_pregunta=request.POST['id_pregunta']),
                           id_evaluacion=Evaluacion.objects.get(
                               id_evaluacion=request.POST['id_evaluacion'])).save()
            except:
                Respuestas.objects.filter(
                    id_pregunta=Preguntas.objects.get(id_pregunta=request.POST['id_pregunta']),
                    id_evaluacion=Evaluacion.objects.get(
                        id_evaluacion=request.POST['id_evaluacion'])).update(
                    respuesta=request.POST['respuesta'])
            if int_siguiente_pregunta == id_preguntas[-1] or 'salir' in request.POST:
                return redirect('evaluacion_ver_preguntas', id_evaluacion=id_evaluacion,
                                id_empleado=id_empleado)
            elif 'guardar' in request.POST:
                # Redireccionamos a la siguiente pregunta
                id_preguntaBuscada = buscar_siguiente_pregunta(id_preguntas, int_siguiente_pregunta)
                return redirect('evaluacionRespuestas', id_evaluacion=id_evaluacion,
                                id_empleado=id_empleado,
                                siguiente_pregunta=id_preguntaBuscada)
    else:
        return render(request, '404.html')


@login_required(login_url='/accounts/login/')
@group_required('admin', 'empleado')
def evaluacion_ver_preguntas(request, id_evaluacion, id_empleado):
    log_access(request)
    comprobacion_emp = comprobar_usuario(request, id_empleado)
    comprobar_resp = comprobar_responsable(request, id_evaluacion)
    if comprobacion_emp or comprobar_resp:
        usuario = request.user.username
        actividad = Actividades(usuario=usuario, accion='CONSULTA EVALUACION',
                                informacion=id_evaluacion)
        actividad.save()
        context = dict()
        preguntas_contestada = dict()
        preguntas_contestada_anterior = dict()
        lista_preguntas_respuestas = list()
        lista_preguntas_respuestas_anterior = list()
        datosEvaluacion = Evaluacion.objects.get(id_evaluacion=id_evaluacion,
                                                 empleado=Empleados.objects.get(id_empleado=id_empleado))
        fecha = datosEvaluacion.anio
       # Datos del empleado de la evaluacion
        context['id_empleado_find'] = datosEvaluacion.empleado.id_empleado
        context['nombre_find'] = datosEvaluacion.empleado.nombre + ' ' + datosEvaluacion.empleado.apellidos
        context['depart_find'] = datosEvaluacion.empleado.area_empleado
        context['rol'] = datosEvaluacion.empleado.rol
        context['identificador_find'] = datosEvaluacion.empleado.identificador
        context['email'] = datosEvaluacion.empleado.email
        context['antiguedad_find'] = datosEvaluacion.empleado.f_antiguedad
        context['responsable_find'] = datosEvaluacion.responsable.nombre + ' ' + datosEvaluacion.responsable.apellidos

        secciones = Secciones.objects.all()

        # Datos empleado logueado
        if request.user.username == 'admin':
            empleado_log = Empleados.objects.get(id_empleado=id_empleado)

        else:
            empleado_log = Empleados.objects.get(email=request.user.username)

        context['id_empleado'] = empleado_log.id_empleado
        context['name_menu'] = empleado_log.nombre
        context['identificador'] = empleado_log.identificador

        # Comprobar permisos de visualizacion
        if empleado_log.evaluador or empleado_log.resp_depart or empleado_log.email == 'admin':
            context['menu'] = True
        if empleado_log.ver_sueldos:
            context['ver_sueldo'] = True


        if datosEvaluacion.id_plantilla.autoevaluacion:
            context['titulo'] = "AUTO EVALUACION"
            if empleado_log.id_empleado == datosEvaluacion.empleado.id_empleado:
                context['vistaEmp'] = True
        elif datosEvaluacion.id_plantilla.de_responsable:
            context['titulo'] = "EVALUACION AL RESPONSABLE"
            if empleado_log.id_empleado == datosEvaluacion.empleado.id_empleado:
                context['evalResp'] = False
                context['vistaEmp'] = True
        else:
            context['titulo'] = "EVALUACION DE DESEMPEÑO"
            if datosEvaluacion.responsable == empleado_log or empleado_log.email == 'admin':
                context['permiso'] = True
        context['valida'] = "Al validar las respuestas...."
        context['verificado'] = datosEvaluacion.verificado
        context['id_evaluacion'] = id_evaluacion

        if 'RESPONSABLE' not in context['titulo']:
            context['grafico'] = chartEmpleado(id_empleado)

        try:
            observaciones = datosEvaluacion.observaciones
            if observaciones is None:
                context['txtobservaciones'] = ""
            else:
                context['txtobservaciones'] = observaciones
        except:
            context['txtobservaciones'] = ""

        if request.method == "GET":

                lista_id_anterior = list(Evaluacion.objects.filter(empleado_id=id_empleado,
                                                                   anio=fecha - 1).values_list('id_evaluacion',
                                                                                                    flat=True))
                for seccion in secciones:
                    preguntas = Preguntas.objects.filter(id_plantilla=datosEvaluacion.id_plantilla, id_seccion=seccion)

                    for pregunta in preguntas:
                        try:
                            rsp = Respuestas.objects.get(id_pregunta=pregunta.id_pregunta, id_evaluacion=id_evaluacion)
                            rsp.respuesta = rsp.respuesta.replace('\r\n', ' ')
                            lista_preguntas_respuestas.append((pregunta, rsp, 'MODIFICAR'))
                            preguntas_contestada[seccion.descripcion] = lista_preguntas_respuestas
                        except Respuestas.DoesNotExist:
                            lista_preguntas_respuestas.append((pregunta, " ", 'RESPONDER'))
                            preguntas_contestada[seccion.descripcion] = lista_preguntas_respuestas
                    lista_preguntas_respuestas = list()
                context['preguntas_contestada'] = preguntas_contestada

                    #reunimos todas las respuestas de todas las evaluaciones del año anterior
                if (len(lista_id_anterior)!=0):
                    for seccion in secciones:
                        preguntas_anteriores = Preguntas.objects.filter(id_seccion=seccion)
                        lista_preguntas_respuestas_anterior = list()
                        for pregunta in preguntas_anteriores:
                            lista_preguntas_respuestas_anterior.append(pregunta)
                            for id in lista_id_anterior:
                                try:
                                    rsp_anterior = Respuestas.objects.get(id_pregunta=pregunta.id_pregunta,
                                                                          id_evaluacion=id)
                                except Respuestas.DoesNotExist:
                                    rsp_anterior = "No fue contestada el año pasado"
                                finally:
                                    lista_preguntas_respuestas_anterior.append((lista_preguntas_respuestas_anterior, rsp_anterior))

                            preguntas_contestada_anterior[seccion.descripcion] = lista_preguntas_respuestas_anterior
                else:
                    preguntas_contestada_anterior = "\r\n "

                context['preguntas_contestada_anterior'] = preguntas_contestada_anterior
                return render(request, 'evaluaciones/evaluacion_ver_preguntas.html', context)
        elif request.method == 'POST':
            if 'enviar_OBS' in request.POST:
                Evaluacion.objects.filter(id_evaluacion=id_evaluacion).update(
                    observaciones=request.POST['obs'])
                usuario = request.user.username
                actividad = Actividades(usuario=usuario, accion='AÑADIR OBSERVACION',
                                        informacion=id_evaluacion)
                actividad.save()
                return redirect('evaluacion_ver_preguntas', id_evaluacion=id_evaluacion, id_empleado=id_empleado)
            if 'respuesta' in request.POST:
                try:
                    Respuestas(respuesta=request.POST['respuesta'],
                               id_pregunta=Preguntas.objects.get(id_pregunta=request.POST['modal_id_pregunta']),
                               id_evaluacion=Evaluacion.objects.get(
                                   id_evaluacion=id_evaluacion)).save()
                except:
                    Respuestas.objects.filter(
                        id_pregunta=Preguntas.objects.get(id_pregunta=request.POST['modal_id_pregunta']),
                        id_evaluacion=Evaluacion.objects.get(
                            id_evaluacion=id_evaluacion)).update(
                        respuesta=request.POST['respuesta'])
                usuario = request.user.username
                actividad = Actividades(usuario=usuario, accion='AÑADIR RESPUESTA',
                                        informacion='Evaluacion:' + str(id_evaluacion) + ' Respuesta:'
                                                    + str(request.POST['modal_id_pregunta']))
                actividad.save()
                return redirect('evaluacion_ver_preguntas', id_evaluacion=id_evaluacion, id_empleado=id_empleado)
    else:
        return render(request, '404.html')


@login_required(login_url='/accounts/login/')
@group_required('admin', 'empleado')
def editar_respuesta(request, id_evaluacion, id_empleado):
    #modificar respuestas de una evaluacion ya verificada, se regsitra su modificacion en la fecha actualizacion
    log_access(request)
    comprobacion_emp = comprobar_usuario(request, id_empleado)
    comprobar_resp = comprobar_responsable(request, id_evaluacion)
    fecha = datetime.now()
    if comprobacion_emp or comprobar_resp:
        context = dict()
        preguntas_contestada = dict()
        preguntas_contestada_anterior = dict()
        lista_preguntas_respuestas = list()
        lista_preguntas_respuestas_anterior = list()
        datosEvaluacion = Evaluacion.objects.get(id_evaluacion=id_evaluacion,
                                                 empleado=Empleados.objects.get(id_empleado=id_empleado))
        datosEvaluacion.updated=fecha
        datosEvaluacion.save()
        fecha = datosEvaluacion.anio
        # Datos del empleado de la evaluacion
        context['id_empleado_find'] = datosEvaluacion.empleado.id_empleado
        context['nombre_find'] = datosEvaluacion.empleado.nombre + ' ' + datosEvaluacion.empleado.apellidos
        context['depart_find'] = datosEvaluacion.empleado.area_empleado
        context['rol'] = datosEvaluacion.empleado.rol
        context['identificador_find'] = datosEvaluacion.empleado.identificador
        context['email'] = datosEvaluacion.empleado.email
        context['antiguedad_find'] = datosEvaluacion.empleado.f_antiguedad
        context['responsable_find'] = datosEvaluacion.responsable.nombre + ' ' + datosEvaluacion.responsable.apellidos

        secciones = Secciones.objects.all()

        # Datos empleado logueado
        if request.user.username == 'admin':
            empleado_log = Empleados.objects.get(id_empleado=id_empleado)

        else:
            empleado_log = Empleados.objects.get(email=request.user.username)

        context['id_empleado'] = empleado_log.id_empleado
        context['name_menu'] = empleado_log.nombre
        context['identificador'] = empleado_log.identificador

        # Comprobar permisos de visualizacion
        if empleado_log.evaluador or empleado_log.resp_depart or empleado_log.email == 'admin':
            context['menu'] = True
        if empleado_log.ver_sueldos:
            context['ver_sueldo'] = True

        if datosEvaluacion.id_plantilla.autoevaluacion:
            context['titulo'] = "AUTO EVALUACION"
            if empleado_log.id_empleado == datosEvaluacion.empleado.id_empleado:
                context['vistaEmp'] = True
        elif datosEvaluacion.id_plantilla.de_responsable:
            context['titulo'] = "EVALUACION AL RESPONSABLE"
            if empleado_log.id_empleado == datosEvaluacion.empleado.id_empleado:
                context['evalResp'] = False
                context['vistaEmp'] = True
        else:
            context['titulo'] = "EVALUACION DE DESEMPEÑO"
            if datosEvaluacion.responsable == empleado_log or empleado_log.email == 'admin':
                context['permiso'] = True
        #cambiamos a no verificado para que modifique su respuesta
        datosEvaluacion.verificado = False
        datosEvaluacion.save()
        context['valida'] = "Al validar las respuestas...."
        context['verificado'] = datosEvaluacion.verificado
        context['id_evaluacion'] = id_evaluacion
        usuario = request.user.username
        actividad = Actividades(usuario=usuario, accion='MODIFICAR EVALUACION VERIFICADA',
                                informacion=id_evaluacion)
        actividad.save()
        if 'RESPONSABLE' not in context['titulo']:
            context['grafico'] = chartEmpleado(id_empleado)

        try:
            observaciones = datosEvaluacion.observaciones
            if observaciones is None:
                context['txtobservaciones'] = ""
            else:
                context['txtobservaciones'] = observaciones
        except:
            context['txtobservaciones'] = ""

        if request.method == "GET":
            lista_id_anterior = list(Evaluacion.objects.filter(empleado_id=id_empleado,
                                                               id_plantilla=datosEvaluacion.id_plantilla,
                                                               anio=int(fecha) - 1).values_list('id_evaluacion', flat=True))
            for seccion in secciones:
                preguntas = Preguntas.objects.filter(id_plantilla=datosEvaluacion.id_plantilla, id_seccion=seccion)
                for pregunta in preguntas:
                    try:
                        rsp = Respuestas.objects.get(id_pregunta=pregunta.id_pregunta, id_evaluacion=id_evaluacion)
                        rsp.respuesta = rsp.respuesta.replace('\r\n', ' ')
                        lista_preguntas_respuestas.append((pregunta, rsp, 'MODIFICAR'))
                        preguntas_contestada[seccion.descripcion] = lista_preguntas_respuestas
                    except Respuestas.DoesNotExist:
                        lista_preguntas_respuestas.append((pregunta, " ", 'RESPONDER'))
                        preguntas_contestada[seccion.descripcion] = lista_preguntas_respuestas

                        # reunimos todas las respuestas de todas las evaluaciones del año anterior
                    if (len(lista_id_anterior) != 0):
                        for id in lista_id_anterior:
                            try:
                                rsp_anterior = Respuestas.objects.get(id_pregunta=pregunta.id_pregunta,
                                                                          id_evaluacion=id)

                            except Respuestas.DoesNotExist:
                                rsp_anterior = "No fue contestada el año pasado"
                        lista_preguntas_respuestas_anterior.append((pregunta, rsp_anterior))
                        preguntas_contestada_anterior[seccion.descripcion] = lista_preguntas_respuestas_anterior
                    else:
                        preguntas_contestada_anterior = ""

                lista_preguntas_respuestas_anterior = list()
                lista_preguntas_respuestas = list()

            context['preguntas_contestada'] = preguntas_contestada
            context['preguntas_contestada_anterior'] = preguntas_contestada_anterior
            return render(request, 'evaluaciones/evaluacion_ver_preguntas.html', context)
        elif request.method == 'POST':
            if 'enviar_OBS' in request.POST:
                Evaluacion.objects.filter(id_evaluacion=id_evaluacion).update(observaciones=request.POST['obs'])
                usuario = request.user.username
                actividad = Actividades(usuario=usuario, accion='AÑADIR OBSERVACION',
                                        informacion=id_evaluacion)
                actividad.save()
                return redirect('evaluacion_ver_preguntas', id_evaluacion=id_evaluacion, id_empleado=id_empleado)
            if 'respuesta' in request.POST:
                usuario = request.user.username
                actividad = Actividades(usuario=usuario, accion='AÑADIR RESPUESTA',
                                        informacion='Evaluacion: ' +str(id_evaluacion) + ' Respuesta: '
                                                    +str(request.POST['modal_id_pregunta']))
                actividad.save()
                try:
                    Respuestas(respuesta=request.POST['respuesta'],
                                id_pregunta=Preguntas.objects.get(id_pregunta=request.POST['modal_id_pregunta']),
                                id_evaluacion=Evaluacion.objects.get(
                                    id_evaluacion=id_evaluacion)).save()
                except:
                    Respuestas.objects.filter(
                        id_pregunta=Preguntas.objects.get(id_pregunta=request.POST['modal_id_pregunta']),
                        id_evaluacion=Evaluacion.objects.get(id_evaluacion=id_evaluacion)).update(
                        respuesta=request.POST['respuesta'])

            return redirect('evaluacion_ver_preguntas', id_evaluacion=id_evaluacion, id_empleado=id_empleado)
    else:
        return render(request, '404.html')


@login_required(login_url='/accounts/login/')
@group_required('admin', 'responsable')
def empleados_responsable(request, id_empleado):
    #return HttpResponse('Entra aquí')
    comprobacion = comprobar_usuario(request, id_empleado)
    next_year = datetime.now().year + 1
    if comprobacion:
        context = dict()
        empleados = dict()
        datos_empleados = list()
        conocimientos_select = list()
        evaluacion_dict = dict()
        datos_evaluacion = list()
        graficos_empleados = dict()
        empleado_buscado = Empleados.objects.get(id_empleado=id_empleado)
        usuario = request.user.username
        actividad = Actividades(usuario=usuario, accion='CONSULTA EMPLEADOS DEL RESPONSABLE',
                                informacion='')
        actividad.save()
        
        empleado_log = Empleados.objects.get(email=request.user.username)
        # Comprobacion de permisos de visualizacion
        if empleado_log.evaluador or empleado_log.resp_depart or empleado_log.email == 'admin':
            context['permiso'] = True
            context['menu'] = True
        if empleado_log.ver_sueldos:
            context['ver_sueldo'] = True
        if empleado_log.resp_depart:
            mis_empleados = Empleados.objects.filter(area_empleado=empleado_buscado.area_empleado)
        else:
            mis_empleados = Empleados.objects.filter(responsable=empleado_buscado)
            if not mis_empleados:
                context['resp_proyecto'] = True
                evaluaciones_proyecto = Evaluacion.objects.filter(responsable=empleado_log)
                id_empleadosEval = list()
                for evaluacion_proyecto in evaluaciones_proyecto:
                    id_empleadosEval.append(evaluacion_proyecto.empleado.id_empleado)
                mis_empleados = Empleados.objects.filter(id_empleado__in=id_empleadosEval)
        # Datos empleado log
        context['id_empleado'] = empleado_buscado.id_empleado
        context['identificador'] = empleado_buscado.identificador
        context['email'] = empleado_buscado.email
        context['nombre'] = empleado_buscado.nombre + ' ' + empleado_buscado.apellidos
        context['name_menu'] = empleado_buscado.nombre
        context['antiguedad'] = str(empleado_buscado.f_antiguedad)

        for empleado in mis_empleados:
            datos_empleados.append(empleado.nombre + " " + empleado.apellidos)
            datos_empleados.append(empleado.email)
            empleados[empleado.id_empleado] = datos_empleados
            # Grafico sueldos
            graficos_empleados[empleado.identificador] = chartEmpleado(empleado.id_empleado)
            datos_empleados = list()
        # context['graficos'] = graficos_empleados
        context['mis_empleados'] = empleados

        if request.method == 'GET':
            return render(request, 'empleado/resp_empleados.html', context)
        elif request.is_ajax():
            if 'selected' in request.POST.get('event', None):
                usuario = request.user.username
                actividad = Actividades(usuario=usuario, accion='DETALLE EMPLEADO DE RESPONSABLE',
                                        informacion=request.POST['identificadorPost'])
                actividad.save()
                #empleado_select = Empleados.objects.get(id_empleado=request.POST['identificadorPost'])
                empleado_select = Empleados.objects.get(identificador=request.POST['identificadorPost'])
                context['nombre_select'] = empleado_select.nombre
                context['apellidos_select'] = empleado_select.apellidos
                context['identificador_select'] = empleado_select.identificador
                context['departamento_select'] = empleado_select.area_empleado
                context['antiguedad_select'] = str(empleado_select.f_antiguedad)

                conocimientos = Conocimientos.objects.filter(id_empleado=empleado_select.id_empleado)
                for conocimiento in conocimientos:
                    conocimientos_select.append(conocimiento.descripcion)
                context['conocimientos_select'] = conocimientos_select

                try:
                    p = Sueldos_propuestos.objects.get(empleado=empleado_select)
                    #al enviarse por JSON, lo envio 1 x 1
                    if p:
                        context['total'] = intcomma(p.total)
                        context['rf'] = intcomma(p.retribucion_fija)
                        context['vi'] = intcomma(p.varibale_individual)
                        context['ve'] = intcomma(p.varibale_empresa)
                        context['g'] = intcomma(p.guardias)
                        context['kms'] = intcomma(p.kilometros)
                        context['bd'] = intcomma(p.bonus_dietas)
                        context['bs'] = intcomma(p.beneficios_sociales)
                except:
                    context['total'] = "Sin datos"
                    context['rf'] = "Sin datos"
                    context['vi'] = "Sin datos"
                    context['ve'] = "Sin datos"
                    context['g'] = "Sin datos"
                    context['kms'] = "Sin datos"
                    context['bd'] = "Sin datos"
                    context['bs'] = "Sin datos"

                plantilla_eval = PlantillaEvaluacion.objects.filter(de_responsable=False)
                evaluaciones = Evaluacion.objects.filter(empleado=empleado_select,
                                                         id_plantilla__in=plantilla_eval).order_by('-anio')
                for evaluacion in evaluaciones:
                    if empleado_log.resp_depart or evaluacion.responsable == empleado_log or empleado_log.nombre == 'admin':
                        datos_evaluacion.append(evaluacion.anio)
                        if evaluacion.id_plantilla.autoevaluacion:
                            datos_evaluacion.append('Auto evaluacion')
                        else:
                            datos_evaluacion.append('Evaluacion')
                        if evaluacion.verificado:
                            datos_evaluacion.append('Verificado')
                        else:
                            datos_evaluacion.append('No verificado')
                        enlace = "/evaluacion/preguntas/lista/" + str(evaluacion.id_evaluacion) + "/" + str(
                            evaluacion.empleado.id_empleado)
                        datos_evaluacion.append(enlace)
                        evaluacion_dict[evaluacion.id_evaluacion] = datos_evaluacion
                        datos_evaluacion = list()
                context['graficos'] = graficos_empleados[str(empleado_select.identificador)]
                context['evaluacion_select'] = evaluacion_dict
                return JsonResponse(context)
    else:
        return render(request, '404.html')


@login_required(login_url='/accounts/login/')
def change_password(request):
    log_access(request)
    context = dict()
    context['titulo'] = "Cambio de contraseña"
    empleado = Empleados.objects.get(email=request.user.email)
    context['identificador'] = empleado.id_empleado
    context['id_empleado'] = empleado.id_empleado
    context['name_menu'] = empleado.nombre

    if empleado.evaluador or empleado.resp_depart or empleado.email == 'admin':
        context['menu'] = True

    if request.method == 'POST':
        usuario = request.user.username
        actividad = Actividades(usuario=usuario, accion='CAMBIO CONTRASEÑA',
                                informacion='')
        actividad.save()
        form = PasswordChangeForm(request.user, request.POST)
        if form.is_valid():
            user = form.save()
            update_session_auth_hash(request, user)  # Important!
            messages.success(request, 'Contraseña cambiada!')
            return render(request, 'change_password.html', context)
        else:
            messages.error(request, 'Please correct the error below.')
    else:
        form = PasswordChangeForm(empleado.id_empleado)
    context['form'] = form
    return render(request, 'change_password.html', context)


@login_required(login_url='/accounts/login/')
@group_required('admin')
def crear_plantilla_responsable(request):
    log_access(request)
    context = dict()
    context['active_plantillas'] = 'class="active"'

    if request.method == 'POST':
        logger.debug('%s - datos recogidos' % request.path)
        PlantillaEvaluacion(descripcion=request.POST.get('descripcion', None),
                            de_responsable=True).save()
        context['title'] = 'Nueva plantilla creada'
        context['tipo'] = "evaluacion"
        context['descripcion'] = request.POST.get('descripcion', None)
        context['autoevaluacion'] = 'No'
        return render(request, 'plantillas/nueva_plantilla_creada.html', context)
    else:
        logger.debug('%s - carga' % request.path)
        context['title'] = 'Nueva plantilla'
        context['form'] = NuevaPlantillaForm
        context['table'] = PlantillasTable(
            data=PlantillaEvaluacion.objects.filter(de_responsable=True).order_by('descripcion'))

        return render(request, 'plantillas/plantilla_responsable.html', context)


@login_required(login_url='/accounts/login/')
@group_required('admin')
def evaluaciones_responsables(request):
    log_access(request)
    context = dict()
    context['title'] = 'Evaluaciones al responsable'
    if request.method == 'GET':
        logger.debug('%s - carga de información' % request.path)
        list_ids_plantilla = PlantillaEvaluacion.objects.filter(autoevaluacion=False, de_responsable=True).values_list(
            'id_plantilla',
            flat=True)
        empleado = Empleados.objects.filter(eliminado=False)  # Coger empleados sin eliminar
        evaluacion = Evaluacion.objects.filter(id_plantilla__in=list_ids_plantilla, empleado__in=empleado)
        table = EvaluacionTable(
            data=evaluacion.order_by('anio'))
        table.paginate(page=request.GET.get('page', 1), per_page=PER_PAGE)
        context['table'] = table
    RequestConfig(request).configure(table)  # to sort
    return render(request, 'evaluaciones/tabla_evaluaciones_responsable.html', context)


@login_required(login_url='/accounts/login/')
@group_required('admin')
def nueva_responsable_evaluacion(request):
    log_access(request)
    context = dict()
    context['title'] = 'Nueva evaluacion'

    if request.method == 'GET':
        logger.debug('%s - carga de información' % request.path)
        context['form'] = NuevaEvaluacionResponsableForm
        return render(request, 'evaluaciones/nueva_evaluacion.html', context)
    else:
        id_plantilla = request.POST.get('id_plantilla', None)
        id_responsable = request.POST.get('responsable', None)
        id_empleado = request.POST.get('empleado', None)
        if id_plantilla and id_responsable and id_empleado:
            plantilla = PlantillaEvaluacion.objects.get(id_plantilla=id_plantilla)
            responsable = Empleados.objects.get(id_empleado=id_responsable)
            empleado = Empleados.objects.get(id_empleado=id_empleado)

            Evaluacion(id_plantilla=plantilla,
                       anio=request.POST.get('anio', None),
                       detalle=request.POST.get('detalle', None),
                       responsable=responsable,
                       empleado=empleado).save()

            context['plantilla'] = plantilla
            context['anio'] = request.POST.get('anio', None)
            context['responsable'] = responsable
            context['empleado'] = empleado
            preguntas = Preguntas.objects.filter(id_plantilla=plantilla.id_plantilla).order_by('updated')
            context['preguntas'] = [pregunta.pregunta for pregunta in preguntas]

        return render(request, 'evaluaciones/nueva_evaluacion_creada.html', context)


@login_required(login_url='/accounts/login/')
@group_required('admin')
def actualizar_sueldos(request, id_empleado):
    #vista para migrar un sueldo propuesto a sueldos, se elimina de propuesto y se añade en normal
    log_access(request)
    context = dict()
    context['id_empleado'] = id_empleado
    context['title'] = 'Actualizar sueldos'
    next_year = datetime.now().year + 1
    empleado = Empleados.objects.get(id_empleado=id_empleado)
    data = Sueldos_propuestos.objects.get(empleado=empleado)
    try:
        #si existe en sueldos actuales ya uno de next_year, se actualiza
        data2 = Sueldos.objects.get(empleado=empleado, anio=int(next_year))
        if data2:
            data2.retribucion_fija = data.retribucion_fija
            data2.varibale_individual = data.varibale_individual
            data2.varibale_empresa = data.varibale_empresa
            data2.beneficios_sociales = data.beneficios_sociales
            data2.kilometros = data.kilometros
            data2.bonus_dietas = data.bonus_dietas
            data2.guardias = data.guardias
            data2.total = data.total
            data2.incremento = data.incremento
            data2.updated = datetime.now()
            (data2).save()

    except:
        #si no existia, se añade
        Sueldos(empleado=data.empleado,
                anio=next_year,
                retribucion_fija=data.retribucion_fija,
                varibale_individual=data.varibale_individual,
                varibale_empresa=data.varibale_empresa,
                beneficios_sociales=data.beneficios_sociales,
                kilometros=data.kilometros,
                bonus_dietas=data.bonus_dietas,
                guardias=data.guardias,
                tipo='Actual',
                total=data.total,
                incremento=data.incremento,
                updated=datetime.now()).save()
    finally:
        #se borra de propuestos
        data.delete()
        return render(request, 'empleados/actualizar_sueldos.html', context)


@login_required(login_url='/accounts/login/')
@group_required('admin')
def detalle_sueldos(request, id_empleado):
    log_access(request)
    context = dict()
    context['title'] = 'Detalle de los sueldos del empleado'
    empleado = Empleados.objects.get(id_empleado=id_empleado)
    context['id_empleado'] = id_empleado
    context['empleado'] = empleado.nombre + " " + empleado.apellidos
    if request.method == 'GET':
       logger.debug('%s - carga de información' % request.path)
       data_sueldos = SueldosTable(data=Sueldos.objects.filter(empleado=empleado))
       data_sueldos.paginate(page=request.GET.get('page', 1), per_page=PER_PAGE_EMPLEADO)
       context['sueldos'] = data_sueldos

       data_propuesto = Sueldos_propuestosTable(data=Sueldos_propuestos.objects.filter(empleado=empleado))
       data_propuesto.paginate(page=request.GET.get('page', 1), per_page=PER_PAGE_EMPLEADO)
       context['propuesto'] = data_propuesto
       context['id'] = Sueldos_propuestos.objects.get(empleado=empleado)
    #else:
    #   logger.debug('%s - subida archivo' % request.path)

    return render(request, 'empleados/detalle_sueldos.html', context)


@login_required(login_url='/accounts/login/')
@group_required('admin')
def actividades(request):
    #lista de actividades que realizan los usuarios
    #opciones: login, nuevo conoc, descarga pdf/pdf rp/pdf ep, carga foto, consulta progreso evals, consulta sueldos, perfil, actualizar conoc,
    #borrar conoc, consulta eval, añadir obs/respuesta, añadir obs/respuesta externo, modificar eval verificada, lista empleados responsable,
    #detalle empleado, cambio contraseña, compartir eval con externo
    context = dict()
    context['title'] = 'Actividades'
    tabla = ActividadesTable(data=Actividades.objects.all())
    tabla.paginate(page=request.GET.get('page', 1), per_page=PER_PAGE)
    context['tabla_actividades'] = tabla
    return render(request, 'actividades/actividades.html', context)


@login_required(login_url='/accounts/login/')
@group_required('admin')
def responsable_evaluacion(request):
    log_access(request)
    context = dict()
    cont = 0
    evaluaciones = Evaluacion.objects.all()
    responsables = list()
    id_responsable = list()
    datosResponsable = list()
    evaluacionesPorResponsable = dict()
    evaluacionSinVerificar = list()
    context['title'] = "Responsable Evaluaciones"
    if request.method == 'GET':

        for evaluacion in evaluaciones:
            responsables.append(evaluacion.responsable)  # Guardo todos los responsables en una lista
            if not evaluacion.verificado:
                evaluacionSinVerificar.append(evaluacion)  # Evaluaciones sin verificar
        evaluacionResponsable = set(responsables)  # Guardar los responsables sin repetidos
        for responsable in evaluacionResponsable:
            datosResponsable.append(responsables.count(responsable))  # Contar cuantas evaluaciones tiene el responsable
            id_responsable.append(responsable.id_empleado)
            for evaluacion in evaluacionSinVerificar:
                if evaluacion.responsable == responsable:
                    cont += 1
            datosResponsable.append(cont)
            evaluacionesPorResponsable[responsable] = datosResponsable
            datosResponsable = []
            cont = 0
        context['id_responsable'] = id_responsable
        context['responsable'] = evaluacionesPorResponsable
    return render(request, 'evaluaciones/responsable_evaluacion.html', context)


@login_required(login_url='/accounts/login/')
@group_required('responsable', 'admin', 'empleado')
def compartir_link(request, id_empleado, id_evaluacion):
    context = dict()
    datosEvaluacion = Evaluacion.objects.get(id_evaluacion=id_evaluacion,
                                             empleado=Empleados.objects.get(id_empleado=id_empleado))
    if datosEvaluacion.id_plantilla.autoevaluacion:
        context['titulo'] = "AUTO EVALUACION"
        evaluador = Empleados.objects.get(id_empleado=id_empleado)
        #evaluador.id_empleado = id_empleado
        evaluado = evaluador
    elif datosEvaluacion.id_plantilla.de_responsable:
        context['titulo'] = "EVALUACION AL RESPONSABLE"
        evaluador = Empleados.objects.get(id_empleado=id_empleado)
        evaluado = evaluador.responsable
    else:
        context['titulo'] = "EVALUACION DE DESEMPEÑO"
        evaluado = Empleados.objects.get(id_empleado=id_empleado)
        evaluador = evaluado.responsable
    #se genera un mail al externo. Cuidado con el evaluador y el evaluado en cada caso
    if request.method == "GET":
        context['form'] = PasswordResetForm()
        return render(request, 'externo/compartir_link.html', context)
    if request.method == "POST":
        form = PasswordResetForm(request.POST)
        if form.is_valid():
            mail = form.cleaned_data['email'],
            correo = form.cleaned_data['email']
            subject = "Evaluación compartida"
            email_template_name = "externo/texto_compartir_externo.html"
            #token valido hasta que el externo valide su evaluacion, luego cambia
            token = (uuid.uuid4().time_low)
            Tokens(token=token,
                   id_empleado=id_empleado,
                   email=correo,
                   id_evaluacion=id_evaluacion,
                   estado='Link compartido').save()
            c = {
                'user': evaluador,
                'evaluado': evaluado,
                'domain':  request.scheme + '://' + request.META['HTTP_HOST'] + '/externo/confirmacion/',
                'var': token,
                'id_evaluacion': id_evaluacion,
                'id_empleado': id_empleado,
            }
            email = render_to_string(email_template_name, c)
            try:
                send_mail(subject, email, '', mail, fail_silently=False)
            except BadHeaderError:
                return HttpResponse('Invalid header found.')
            finally:
                usuario = request.user.username
                actividad = Actividades(usuario=usuario, accion='COMPARTIR EVALUACION CON EXTERNO',
                                        informacion='Evaluacion: ' + str(id_evaluacion) + ' Externo: ' + correo)
                actividad.save()
            return redirect('enviar_email', evaluador.id_empleado)
        form = PasswordResetForm()
        return render(request=request, template_name="externo/compartir_link.html",
                      context={"form": form})


@login_required(login_url='/accounts/login/')
@group_required('admin', 'responsable', 'empleado')
def enviar_email(request, id_empleado):
    context = dict()
    context['perfil'] = '/empleado/perfil/' + str(id_empleado)
    return render(request, 'externo/enviar_email.html', context)


@anonymous_required()
def confirmacion(request, var, id_evaluacion, id_empleado):
    context = dict()
    context['var'] = var
    context['id_evaluacion'] = id_evaluacion
    context['id_empleado'] = id_empleado #evaluado
    fecha_actual = date.today()
    try:
        #periodo maximo de 7 dias y comprobar correo del externo y del evaluado
        variable = Tokens.objects.get(token=var, id_evaluacion=id_evaluacion, id_empleado=id_empleado)
        fecha_maxima = variable.updated + timedelta(days=RESET_TIMEOUT_DAYS)
        if variable and fecha_maxima >= fecha_actual:
            try:
                empleado = Empleados.objects.get(
                    id_empleado=id_empleado)  # evaluado
                correo = empleado.email  # correo del evaluado interno de indizen
            finally:
                if request.method == 'POST':
                    #puede introducirse pepito@indizen.com o solo pepito
                    if ('mail_field' in request.POST) and ('mail_empleado' in request.POST):
                        if (request.POST['mail_field'].strip() == variable.email) and \
                                ((request.POST['mail_empleado'].strip() + EMAIL_SUFIX == correo) or
                                 (request.POST['mail_empleado'].strip() == correo)):
                            return redirect('rellenar_evaluacion', var, id_evaluacion, id_empleado)
                        else:
                            return render(request, '404_externo.html')
                    else:
                        context['Title'] = 'Confirma tu email'
                        return render(request, 'externo/confirmacion.html', context)
                elif request.method == 'GET':
                    context['Title'] = 'Confirma tu email'
                    return render(request, 'externo/confirmacion.html', context)
        elif fecha_maxima < fecha_actual:
            update = Tokens.objects.get(token=var, id_evaluacion=id_evaluacion)
            update.token = (uuid.uuid4().time_low)
            update.estado = 'Link caducado'
            update.save()
            return render(request, '404_externo.html')
        else:
            return render(request, '404_externo.html')
    except:
        return render(request, '404_externo.html')



@anonymous_required()
def rellenar_evaluacion(request, var, id_evaluacion, id_empleado):
    #no se necesita usuario
    variable = Tokens.objects.filter(token=var, id_evaluacion=id_evaluacion, id_empleado=id_empleado)
    if variable:
        variable = Tokens.objects.get(token=var, id_evaluacion=id_evaluacion, id_empleado=id_empleado)
        context = dict()
        context['var'] = var
        context['id_evaluacion'] = id_evaluacion
        context['id_empleado'] = id_empleado
        preguntas_contestada = dict()
        lista_preguntas_respuestas = list()
        datosEvaluacion = Evaluacion.objects.get(id_evaluacion=id_evaluacion,
                                                 empleado=Empleados.objects.get(id_empleado=id_empleado))
        context['id_empleado'] = id_empleado
        # Datos del empleado de la evaluacion. EVALUADO
        context['id_empleado_find'] = datosEvaluacion.empleado.id_empleado
        context['nombre_find'] = datosEvaluacion.empleado.nombre + ' ' + datosEvaluacion.empleado.apellidos
        context['depart_find'] = datosEvaluacion.empleado.area_empleado
        context['identificador_find'] = datosEvaluacion.empleado.identificador
        context['email'] = datosEvaluacion.empleado.email
        context['responsable_find'] = datosEvaluacion.responsable.nombre + ' ' + datosEvaluacion.responsable.apellidos

        secciones = Secciones.objects.all()

        if datosEvaluacion.id_plantilla.autoevaluacion:
            context['titulo'] = "AUTO EVALUACION"
            context['evaluador'] = id_empleado
        elif datosEvaluacion.id_plantilla.de_responsable:
            context['titulo'] = "EVALUACION AL RESPONSABLE"
            context['evaluador'] = id_empleado
        else:
            context['titulo'] = "EVALUACION DE DESEMPEÑO"
            evaluado = Empleados.objects.get(id_empleado=id_empleado)
            context['evaluador'] = evaluado.responsable.id_empleado

        context['valida'] = "Al validar las respuestas...."
        context['verificado'] = datosEvaluacion.verificado
        try:
            observaciones = datosEvaluacion.observaciones
            if observaciones is None:
                context['txtobservaciones'] = ""
            else:
                context['txtobservaciones'] = observaciones
        except:
            context['txtobservaciones'] = ""

        try:
            externo = datosEvaluacion.externo
            if externo is None:
                context['externo'] = ""
            else:
                context['externo'] = externo
        except:
            context['externo'] = ""

        if request.method == "GET":
            for seccion in secciones:
                preguntas = Preguntas.objects.filter(id_plantilla=datosEvaluacion.id_plantilla, id_seccion=seccion)
                for pregunta in preguntas:
                    try:
                        rsp = Respuestas.objects.get(id_pregunta=pregunta.id_pregunta, id_evaluacion=id_evaluacion)
                        rsp.respuesta = rsp.respuesta.replace('\r\n', ' ')
                        lista_preguntas_respuestas.append((pregunta, rsp, 'RESPONDER'))
                        preguntas_contestada[seccion.descripcion] = lista_preguntas_respuestas
                    except Respuestas.DoesNotExist:
                        lista_preguntas_respuestas.append((pregunta, " ", 'RESPONDER'))
                        preguntas_contestada[seccion.descripcion] = lista_preguntas_respuestas
                lista_preguntas_respuestas = list()

            context['preguntas_contestada'] = preguntas_contestada
            return render(request, 'externo/evaluacion_externo.html', context)
        elif request.method == 'POST':
            if 'enviar_OBS' in request.POST:
                actividad = Actividades(usuario=variable.email, accion='AÑADIR OBSERVACION POR EXTERNO',
                                        informacion=id_evaluacion)
                actividad.save()
                Evaluacion.objects.filter(id_evaluacion=id_evaluacion).update(observaciones=request.POST['obs'])
                return redirect('rellenar_evaluacion', var, id_evaluacion, id_empleado)
            if 'enviar_NOMBRE' in request.POST:
                #nombre del externo que rellena la evaluacion, no es obligatorio porque queda registrado en tokens
                Evaluacion.objects.filter(id_evaluacion=id_evaluacion).update(externo=request.POST['externo'])
                return redirect('rellenar_evaluacion', var, id_evaluacion, id_empleado)
            if 'respuesta' in request.POST:
                actividad = Actividades(usuario=variable.email, accion='AÑADIR RESPUESTA POR EXTERNO',
                                        informacion='Evaluacion: ' + str(id_evaluacion) + ' Respuesta: '
                                                    + str(request.POST['modal_id_pregunta']))
                actividad.save()
                try:
                    Respuestas(respuesta=request.POST['respuesta'],
                               id_pregunta=Preguntas.objects.get(id_pregunta=request.POST['modal_id_pregunta']),
                               id_evaluacion=Evaluacion.objects.get(id_evaluacion=id_evaluacion)).save()
                except:
                    Respuestas.objects.filter(
                        id_pregunta=Preguntas.objects.get(id_pregunta=request.POST['modal_id_pregunta']),
                        id_evaluacion=Evaluacion.objects.get(
                            id_evaluacion=id_evaluacion)).update(respuesta=request.POST['respuesta'])

                return redirect('rellenar_evaluacion', var, id_evaluacion, id_empleado)
        else:
            return render(request, '404_externo.html')
    else:
        return render(request, '404_externo.html')


@anonymous_required()
def confirmar_externo(request, var, id_evaluacion, id_empleado, evaluador):
    #cambia token una vez externo manda sus respuestas de evaluacion
    update = Tokens.objects.get(token=var, id_evaluacion=id_evaluacion)
    update.token = (uuid.uuid4().time_low)
    update.estado = 'Link correcto usado por externo'
    update.save()
    context = dict()
    context['mensaje'] = 'Confirmado todo bien'
    evaluador_indizen = Empleados.objects.get(id_empleado=evaluador)
    correo = evaluador_indizen.email,
    subject = "Validar evaluacion por externo"
    email_template_name = "externo/texto_validar.html"
    #se envia un email al evaluador para que verifique las respuestas
    c = {
        'user': request.user,
        'domain': request.scheme + '://' + request.META['HTTP_HOST'] + '/evaluacion/preguntas/lista',
        'site_name': 'Indizen',
        'id_empleado': id_empleado,
        'id_evaluacion': id_evaluacion,
    }
    email = render_to_string(email_template_name, c)
    try:
        send_mail(subject, email, '', correo, fail_silently=False)
    except BadHeaderError:
        return HttpResponse('Invalid header found.')
    return render(request, 'externo/confirmar_externo.html')


@login_required(login_url='/accounts/login/')
@group_required('admin')
def exportar_modelosORIGINAL(request, modelo):
    #usar si queremos un CSV como resultado final
    #modelo 1 = empleados, modelo 2 = sueldos, modelo 3 = evaluaciones, modelo 4 = preguntas, modelo 5 = respuestas
    response = HttpResponse(content_type='text/csv')
    writer = csv.writer(response)
    if modelo == 1:
        writer.writerow(['identificador', 'email', 'nombre', 'apellidos', 'f_antiguedad', 'area_empleado', 'rol', 'evaluador', 'responsable_identificador', 'resp_depart', 'ver_sueldos'])
        model = Empleados.objects.all().values_list('identificador','email','nombre','apellidos','f_antiguedad','area_empleado','rol','evaluador','responsable','resp_depart','ver_sueldos')
        name = "database_empleados.csv"
        model2 = False
    elif modelo == 2:
        writer.writerow(['id_empleado','retribucion_fija', 'variable_individual', 'variable_empresa', 'beneficios_sociales', 'bonus_dietas', 'kilometros', 'guardias', 'tipo', 'total', 'incremento', 'anio'])
        model = Sueldos.objects.all().values_list('empleado','retribucion_fija', 'varibale_individual', 'varibale_empresa', 'beneficios_sociales', 'bonus_dietas', 'kilometros', 'guardias', 'tipo', 'total', 'incremento', 'anio')
        model2 = Sueldos_propuestos.objects.all().values_list('empleado','retribucion_fija', 'varibale_individual', 'varibale_empresa', 'beneficios_sociales', 'bonus_dietas', 'kilometros', 'guardias', 'tipo', 'total', 'incremento')
        name = "database_sueldos.csv"
    elif modelo == 3:
        writer.writerow(['plantilla', 'anio', 'detalle', 'responsable_identificador', 'empleado_identificador', 'verificado', 'fecha_verificado', 'observaciones', 'externo'])
        model = Evaluacion.objects.all().values_list('id_plantilla', 'anio', 'detalle', 'responsable', 'empleado', 'verificado', 'fecha_verificado', 'observaciones', 'externo')
        name = "database_evaluaciones.csv"
        model2 = False
    elif modelo == 4:
        writer.writerow(['id_plantilla', 'id_seccion', 'pregunta'])
        model = Preguntas.objects.all().values_list('id_plantilla','id_seccion','pregunta')
        name = "database_preguntas.csv"
        model2 = False
    elif modelo == 5:
        writer.writerow(['id_evaluacion', 'id_pregunta', 'respuesta'])
        model = Respuestas.objects.all().values_list('id_evaluacion','id_pregunta','respuesta')
        name = "database_respuestas.csv"
        model2 = False
    else:
        return render(request, '404.html')

    for data in model:
        writer.writerow(data)
    if model2:
        for data in model2:
            writer.writerow(data)
    #se puede cambiar el nombre segun el modelo igual que esta hecho en exportar_modelos
    response['Content-Disposition'] = 'attachment; filename="database.csv"'
    return response


@login_required(login_url='/accounts/login/')
@group_required('admin')
def exportar_modelos(request, modelo):
    #modelo 1 = empleados, modelo 2 = sueldos, modelo 3 = evaluaciones, modelo 4 = preguntas, modelo 5 = respuestas
    #usar si queremos un EXCEL como resultado final
    w = Workbook()
    writer = w.active
    response = HttpResponse(content_type='application/ms-excel')
    if modelo == 1:
        writer['A1'] = ('identificador')
        writer['B1'] = ('email')
        writer['C1'] = ('nombre')
        writer['D1'] = ('apellidos')
        writer['E1'] = ('f_antiguedad')
        writer['F1'] = ('area_empleado')
        writer['G1'] = ('rol')
        writer['H1'] = ('evaluador')
        writer['I1'] = ('responsable')
        writer['J1'] = ('resp_depart')
        writer['K1'] = ('ver_sueldos')
        model = Empleados.objects.all().values_list('identificador', 'email', 'nombre', 'apellidos', 'f_antiguedad',
                                                    'area_empleado', 'rol', 'evaluador', 'responsable', 'resp_depart',
                                                    'ver_sueldos')
        name = "database_empleados.xlsx"
        count = 2
        #count = numero de fila
        for m in model:
            writer.cell(row=count, column=1).value = m[0]
            writer.cell(row=count, column=2).value = m[1]
            writer.cell(row=count, column=3).value = m[2]
            writer.cell(row=count, column=4).value = m[3]
            writer.cell(row=count, column=5).value = str(m[4])
            writer.cell(row=count, column=6).value = m[5]
            writer.cell(row=count, column=7).value = m[6]
            writer.cell(row=count, column=8).value = str(m[7])
            writer.cell(row=count, column=9).value = m[8]
            writer.cell(row=count, column=10).value = str(m[9])
            writer.cell(row=count, column=11).value = str(m[10])
            count = count+1

    elif modelo == 2:
        writer['A1'] = ('id_empleado')
        writer['B1'] = ('retribucion_fija')
        writer['C1'] = ('variable_individual')
        writer['D1'] = ('variable_empresa')
        writer['E1'] = ('beneficios_sociales')
        writer['F1'] = ('bonus_dietas')
        writer['G1'] = ('kilometros')
        writer['H1'] = ('guardias')
        writer['I1'] = ('tipo')
        writer['J1'] = ('total')
        writer['K1'] = ('incremento')
        writer['L1'] = ('anio')
        model = Sueldos.objects.all().values_list('empleado','retribucion_fija', 'varibale_individual', 'varibale_empresa', 'beneficios_sociales', 'bonus_dietas', 'kilometros', 'guardias', 'tipo', 'total', 'incremento', 'anio')
        model2 = (Sueldos_propuestos.objects.all().values_list('empleado','retribucion_fija', 'varibale_individual', 'varibale_empresa', 'beneficios_sociales', 'bonus_dietas', 'kilometros', 'guardias', 'tipo', 'total', 'incremento'))
        name = "database_sueldos.xlsx"
        count = 2
        for m in model:
            writer.cell(row=count, column=1).value = m[0]
            writer.cell(row=count, column=2).value = m[1]
            writer.cell(row=count, column=3).value = m[2]
            writer.cell(row=count, column=4).value = m[3]
            writer.cell(row=count, column=5).value = (m[4])
            writer.cell(row=count, column=6).value = m[5]
            writer.cell(row=count, column=7).value = m[6]
            writer.cell(row=count, column=8).value = (m[7])
            writer.cell(row=count, column=9).value = str(m[8])
            writer.cell(row=count, column=10).value = (m[9])
            writer.cell(row=count, column=11).value = (m[10])
            writer.cell(row=count, column=11).value = (m[11])
            count = count + 1

        for m in model2:
            writer.cell(row=count, column=1).value = m[0]
            writer.cell(row=count, column=2).value = m[1]
            writer.cell(row=count, column=3).value = m[2]
            writer.cell(row=count, column=4).value = m[3]
            writer.cell(row=count, column=5).value = (m[4])
            writer.cell(row=count, column=6).value = m[5]
            writer.cell(row=count, column=7).value = m[6]
            writer.cell(row=count, column=8).value = (m[7])
            writer.cell(row=count, column=9).value = str(m[8])
            writer.cell(row=count, column=10).value = (m[9])
            writer.cell(row=count, column=11).value = (m[10])
            count = count + 1

    elif modelo == 3:
        writer['A1'] = ('plantilla')
        writer['B1'] = ('anio')
        writer['C1'] = ('detalle')
        writer['D1'] = ('responsable_identificador')
        writer['E1'] = ('empleado_identificador')
        writer['F1'] = ('verificado')
        writer['G1'] = ('fecha_verificado')
        writer['H1'] = ('observaciones')
        writer['I1'] = ('externo')
        model = Evaluacion.objects.all().values_list('id_plantilla', 'anio', 'detalle', 'responsable', 'empleado', 'verificado', 'fecha_verificado', 'observaciones', 'externo')
        name = "database_evaluaciones.xlsx"
        count = 2
        for m in model:
            writer.cell(row=count, column=1).value = m[0]
            writer.cell(row=count, column=2).value = m[1]
            writer.cell(row=count, column=3).value = str(m[2])
            writer.cell(row=count, column=4).value = m[3]
            writer.cell(row=count, column=5).value = (m[4])
            writer.cell(row=count, column=6).value = str(m[5])
            writer.cell(row=count, column=7).value = str(m[6])
            writer.cell(row=count, column=8).value = str(m[7])
            writer.cell(row=count, column=9).value = str(m[8])
            count = count + 1

    elif modelo == 4:
        writer['A1'] = ('id_plantilla')
        writer['B1'] = ('id_seccion')
        writer['C1'] = ('pregunta')
        model = Preguntas.objects.all().values_list('id_plantilla','id_seccion','pregunta')
        name = "database_preguntas.xlsx"
        count = 2
        for m in model:
            writer.cell(row=count, column=1).value = m[0]
            writer.cell(row=count, column=2).value = m[1]
            writer.cell(row=count, column=3).value = str(m[2])
            count = count + 1

    elif modelo == 5:
        writer['A1'] = ('id_evaluacion')
        writer['B1'] = ('id_pregunta')
        writer['C1'] = ('respuesta')
        model = Respuestas.objects.all().values_list('id_evaluacion','id_pregunta','respuesta')
        name = "database_respuestas.xlsx"
        count = 2
        for m in model:
            writer.cell(row=count, column=1).value = m[0]
            writer.cell(row=count, column=2).value = m[1]
            writer.cell(row=count, column=3).value = str(m[2])
            count = count + 1
    else:
        return render(request, '404.html')

    response['Content-Disposition'] = 'attachment; filename={0}'.format(name)
    w.save(response)
    return response
